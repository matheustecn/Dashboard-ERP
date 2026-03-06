import base64
import io
import datetime
import os
import re
import dash
from dash import dcc, html, dash_table, Input, Output, State, ctx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go
import plotly.express as px
import pandas as pd

# ── CONFIGURAÇÃO DO APP ──────────────────────────────────────────────────────
app = dash.Dash(__name__, suppress_callback_exceptions=True)
server = app.server  # required for Railway / Render / Gunicorn

# ── MAPEAMENTO DE CAMPOS DRE ──────────────────────────────────────────────────
DRE_CAMPOS = {
    'venda_vista':  ['venda_vista', 'vendas_a_vista', 'vista', 'venda a vista'],
    'venda_prazo':  ['venda_prazo', 'vendas_a_prazo', 'prazo', 'venda a prazo'],
    'g_bolos':      ['g_bolos', 'bolos', 'bolos_tortas', 'bolos e tortas'],
    'g_buffet':     ['g_buffet', 'buffet'],
    'g_cafes':      ['g_cafes', 'cafes', 'cafes_sucos', 'cafes e sucos'],
    'g_lanches':    ['g_lanches', 'lanches'],
    'g_porcoes':    ['g_porcoes', 'porcoes', 'porções'],
    'g_salgados':   ['g_salgados', 'salgados'],
    'g_drinks':     ['g_drinks', 'drinks'],
    'g_conv':       ['g_conv', 'conveniencia', 'conveniência'],
    'g_cigarros':   ['g_cigarros', 'cigarros'],
    'g_bebidas':    ['g_bebidas', 'bebidas'],
    'g_picole':     ['g_picole', 'picole', 'picolé'],
    'g_estcoz':     ['g_estcoz', 'estoque_cozinha', 'estcoz'],
    'g_bichos':     ['g_bichos', 'bichos', 'bichos_brinquedos'],
    'g_carnes':     ['g_carnes', 'carnes'],
    'or_juros':     ['or_juros', 'juros_recebidos', 'juros recebidos'],
    'or_alug':      ['or_alug', 'alugueis_receita', 'alugueis recebidos'],
    'or_outras':    ['or_outras', 'outras_receitas', 'outras receitas'],
    'or_fundos':    ['or_fundos', 'fundos', 'rendimentos_fundos'],
    'or_bonif':     ['or_bonif', 'bonificacoes', 'bonificações recebidas'],
    'c_bebidas':    ['c_bebidas', 'compra_bebidas', 'compras bebidas'],
    'c_bolos':      ['c_bolos', 'compra_bolos'],
    'c_carnes':     ['c_carnes', 'compra_carnes'],
    'c_cigarros':   ['c_cigarros', 'compra_cigarros'],
    'c_conv':       ['c_conv', 'compra_conveniencia'],
    'c_estcoz':     ['c_estcoz', 'compra_estcoz'],
    'c_picoles':    ['c_picoles', 'compra_picoles'],
    'c_salgados':   ['c_salgados', 'compra_salgados'],
    'c_insumos':    ['c_insumos', 'insumos', 'compra_insumos'],
    'c_embal':      ['c_embal', 'embalagens', 'compra_embalagens'],
    'c_recarga':    ['c_recarga', 'recarga', 'recarga_celular'],
    'c_bichos':     ['c_bichos', 'compra_bichos'],
    'c_buffet':     ['c_buffet', 'compra_buffet'],
    'cf_energia':   ['cf_energia', 'energia', 'energia_eletrica', 'energia elétrica'],
    'cf_tel':       ['cf_tel', 'telefone', 'tel'],
    'cf_agua':      ['cf_agua', 'agua', 'água'],
    'cf_iptu':      ['cf_iptu', 'iptu'],
    'cf_sal':       ['cf_sal', 'salarios', 'salários', 'folha'],
    'cf_enc':       ['cf_enc', 'encargos', 'encargos_sociais'],
    'cf_imp':       ['cf_imp', 'impostos'],
    'cf_seg':       ['cf_seg', 'seguro_vida', 'seguro vida'],
    'cf_diar':      ['cf_diar', 'diarias', 'diárias'],
    'cv_sist':      ['cv_sist', 'sistemas'],
    'cv_terc':      ['cv_terc', 'servicos_terceiros', 'terceiros'],
    'cv_exped':     ['cv_exped', 'expediente'],
    'cv_honor':     ['cv_honor', 'honorarios', 'honorários'],
    'cv_manut':     ['cv_manut', 'manutencao', 'manutenção'],
    'cv_viag':      ['cv_viag', 'viagens'],
    'cv_taxas':     ['cv_taxas', 'taxas'],
    'cv_unif':      ['cv_unif', 'uniformes'],
    'cv_limp':      ['cv_limp', 'limpeza'],
    'cv_alug':      ['cv_alug', 'aluguel', 'alugueis'],
    'cv_caixa':     ['cv_caixa', 'caixa', 'faltas_caixa'],
    'cv_mora':      ['cv_mora', 'juros_mora', 'mora'],
    'cv_banco':     ['cv_banco', 'tarifas_bancarias', 'banco'],
    'cv_cart':      ['cv_cart', 'tarifas_cartao', 'cartao'],
    'cv_brindes':   ['cv_brindes', 'brindes'],
    'cv_utens':     ['cv_utens', 'utensilios', 'utensílios'],
    'cv_veic':      ['cv_veic', 'veiculos', 'veículos'],
    'cv_segpred':   ['cv_segpred', 'seguro_edificacoes'],
    'cv_gerais':    ['cv_gerais', 'despesas_gerais', 'gerais'],
    'cv_gas':       ['cv_gas', 'gas', 'gas_glp'],
    'cv_estoque':   ['cv_estoque', 'contagem_estoque'],
    'cv_comiss':    ['cv_comiss', 'comissoes', 'comissões'],
}

_ALIAS_MAP = {}
for canonical, aliases in DRE_CAMPOS.items():
    for alias in aliases:
        _ALIAS_MAP[alias.lower().strip()] = canonical

def _parse_number(s):
    if not s:
        return None
    s = s.strip().replace('R$', '').replace(' ', '')
    if not s:
        return None
    has_dot   = '.' in s
    has_comma = ',' in s
    if has_dot and has_comma:
        dot_pos   = s.rfind('.')
        comma_pos = s.rfind(',')
        if comma_pos > dot_pos:
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif has_comma and not has_dot:
        after_comma = s.split(',')[-1]
        if len(after_comma) <= 2:
            s = s.replace(',', '.')
        else:
            s = s.replace(',', '')
    try:
        return float(s)
    except ValueError:
        return None

def parse_csv_to_dre(decoded_text):
    lines = [l.strip() for l in decoded_text.strip().split('\n') if l.strip()]
    if not lines:
        return {}, 0
    result = {}
    sep = ';' if lines[0].count(';') >= lines[0].count(',') else ','
    header = [p.strip().strip('"').strip("'").lower() for p in lines[0].split(sep)]
    if len(header) == 2 and any(h in ('campo', 'field', 'conta', 'descricao', 'descrição') for h in header):
        ci = 0 if header[0] in ('campo','field','conta','descricao','descrição') else 1
        vi = 1 - ci
        for line in lines[1:]:
            parts = [p.strip().strip('"').strip("'") for p in line.split(sep)]
            if len(parts) < 2:
                continue
            campo_raw = parts[ci].lower().strip()
            canonical = _ALIAS_MAP.get(campo_raw)
            if not canonical:
                for alias, can in _ALIAS_MAP.items():
                    if alias in campo_raw or campo_raw in alias:
                        canonical = can
                        break
            if not canonical:
                continue
            parsed = _parse_number(parts[vi])
            if parsed is not None:
                result[canonical] = parsed
        return result, len(result)
    canonical_cols = []
    for h in header:
        canonical_cols.append(_ALIAS_MAP.get(h))
    for line in lines[1:]:
        parts = [p.strip().strip('"').strip("'") for p in line.split(sep)]
        for i, val_raw in enumerate(parts):
            if i >= len(canonical_cols) or canonical_cols[i] is None:
                continue
            parsed = _parse_number(val_raw)
            if parsed is not None:
                result[canonical_cols[i]] = parsed
    return result, len(result)


# ══════════════════════════════════════════════════════════════════════════════
# CONCILIAÇÃO BANCÁRIA — FUNÇÕES DE CLASSIFICAÇÃO
# ══════════════════════════════════════════════════════════════════════════════

def classificar_tipo(descricao):
    descricao = str(descricao).upper()
    if "DEVOLUCAO PIX" in descricao or "DEVOLUÇÃO PIX" in descricao:
        return "ESTORNO"
    if "TARIFA COM R LIQUIDACAO" in descricao or "TARIFA COM R LIQUIDAÇÃO" in descricao:
        return "TARIFA BOLETO"
    if "DÉBITO AUTOMÁTICO" in descricao or "DEBITO AUTOMATICO" in descricao:
        return "DEB. AUTOMÁTICO"
    if any(p in descricao for p in ["SERV PIX", "SAQUE", "TROCO"]):
        return "OUTROS VALORES"
    if "DEP DINHEIRO" in descricao or "DEPOSITO" in descricao or "DEPÓSITO" in descricao:
        return "DEPÓSITO"
    if "PAGSEGURO" in descricao:
        return "PAGSEGURO"
    ADQUIRENTES_CARTAO = [
        "CIELO", "ALELO", "LINK CARD", "TRUCKPAG", "STONE", "REDE",
        "PRIME", "NEO", "LINK", "COOPERCARD", "SEM PARAR", "GETNET",
        "SAFRAPAY", "MERCADOPAGO", "SUMUP", "PAGBANK",
    ]
    for adq in ADQUIRENTES_CARTAO:
        if adq in descricao:
            return "CARTÃO"
    if "LIQ.COBRANCA SIMPLES" in descricao or "LIQ COBRANCA" in descricao:
        return "BOLETOS RECEBIDOS"
    if "LIQUIDACAO BOLETO" in descricao or "LIQUIDAÇÃO BOLETO" in descricao or "BOLETOS PAGOS" in descricao:
        return "BOLETOS PAGOS"
    if "PIX" in descricao:
        return "PIX"
    if "TED" in descricao:
        return "TED"
    if "DOC" in descricao:
        return "DOC"
    if "TARIFA" in descricao or "TAR " in descricao:
        return "TARIFA BANCÁRIA"
    if "IOF" in descricao:
        return "IOF"
    if "JUROS" in descricao or "MULTA" in descricao:
        return "JUROS/MULTA"
    return "OUTROS"

def extrair_documento(descricao):
    texto = str(descricao)
    cnpj = re.search(r'\d{14}', texto)
    if cnpj:
        raw = cnpj.group()
        return f"{raw[:2]}.{raw[2:5]}.{raw[5:8]}/{raw[8:12]}-{raw[12:]}"
    cpf = re.search(r'\d{11}', texto)
    if cpf:
        raw = cpf.group()
        return f"{raw[:3]}.{raw[3:6]}.{raw[6:9]}-{raw[9:]}"
    return ""

def classificar_categoria(tipo, valor):
    if valor > 0:
        mapa = {
            "CARTÃO":            "RECEITA CARTÃO",
            "PAGSEGURO":         "RECEITA PAGSEGURO",
            "DEPÓSITO":          "RECEITA DEPÓSITO",
            "BOLETOS RECEBIDOS": "RECEITA BOLETO",
            "PIX":               "RECEITA PIX",
            "TED":               "RECEITA TED",
            "DOC":               "RECEITA DOC",
            "ESTORNO":           "ESTORNO RECEBIDO",
        }
        return mapa.get(tipo, "RECEITA OPERACIONAL")
    else:
        mapa = {
            "BOLETOS PAGOS":    "PAGAMENTO BOLETO",
            "PIX":              "PAGAMENTO PIX",
            "TED":              "PAGAMENTO TED",
            "DOC":              "PAGAMENTO DOC",
            "TARIFA BANCÁRIA":  "DESPESA BANCÁRIA",
            "TARIFA BOLETO":    "DESPESA BANCÁRIA",
            "DEB. AUTOMÁTICO":  "DEB. AUTOMÁTICO",
            "IOF":              "DESPESA BANCÁRIA",
            "JUROS/MULTA":      "JUROS/MULTA",
            "ESTORNO":          "ESTORNO PAGO",
        }
        return mapa.get(tipo, "SAÍDA OPERACIONAL")

def processar_extrato_excel(content_bytes, skiprows=8):
    """
    Lê o Excel do extrato bancário, classifica e retorna um DataFrame tratado.
    Tenta diferentes skiprows caso as colunas esperadas não sejam encontradas.
    """
    COLUNAS_ESPERADAS = ["Data", "Descrição", "Documento", "Valor (R$)", "Saldo (R$)"]
    
    for skip in [skiprows, 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
        try:
            df = pd.read_excel(io.BytesIO(content_bytes), skiprows=skip)
            df.columns = df.columns.str.strip()
            if all(c in df.columns for c in COLUNAS_ESPERADAS):
                break
        except Exception:
            continue
    else:
        # Se não encontrou colunas esperadas, tenta ler e renomear as 5 primeiras colunas
        df = pd.read_excel(io.BytesIO(content_bytes), skiprows=0)
        df.columns = df.columns.str.strip()
        if len(df.columns) >= 5:
            rename_map = {df.columns[0]: "Data", df.columns[1]: "Descrição",
                          df.columns[2]: "Documento", df.columns[3]: "Valor (R$)",
                          df.columns[4]: "Saldo (R$)"}
            df = df.rename(columns=rename_map)
        else:
            raise ValueError("Estrutura do extrato não reconhecida. Verifique o arquivo.")

    # Remove linhas sem data ou valor
    df = df.dropna(subset=["Data", "Valor (R$)"])
    df["Valor (R$)"] = pd.to_numeric(df["Valor (R$)"], errors="coerce")
    df = df.dropna(subset=["Valor (R$)"])

    # Classificações
    df["Tipo"]       = df["Descrição"].apply(classificar_tipo)
    df["Documento_Extraido"] = df["Descrição"].apply(extrair_documento)
    df["Categoria"]  = df.apply(lambda r: classificar_categoria(r["Tipo"], r["Valor (R$)"]), axis=1)
    df["Entrada/Saída"] = df["Valor (R$)"].apply(lambda x: "ENTRADA" if x > 0 else "SAÍDA")

    # Tratamento de Estorno PIX 1:1
    def extrair_id(desc):
        m = re.search(r'\d+', str(desc))
        return m.group() if m else ""
    df["_id_tx"] = df["Descrição"].apply(extrair_id)
    devolucoes = df[df["Descrição"].str.upper().str.contains("DEVOLUCAO PIX|DEVOLUÇÃO PIX", na=False)]
    for idx, row in devolucoes.iterrows():
        id_tx = row["_id_tx"]
        recebimentos = df[
            (df["Descrição"].str.upper().str.contains("RECEBIMENTO PIX", na=False)) &
            (df["_id_tx"] == id_tx) &
            (df["Tipo"] != "ESTORNO")
        ]
        if not recebimentos.empty:
            df.loc[idx, "Tipo"] = "ESTORNO"
            df.loc[recebimentos.index[0], "Tipo"] = "ESTORNO"

    # Formatar data
    try:
        df["Data"] = pd.to_datetime(df["Data"]).dt.strftime("%d/%m/%Y")
    except Exception:
        df["Data"] = df["Data"].astype(str)

    df_final = pd.DataFrame({
        "Data":              df["Data"],
        "Entrada/Saída":     df["Entrada/Saída"],
        "Tipo":              df["Tipo"],
        "Descrição":         df["Descrição"],
        "Documento":         df["Documento_Extraido"],
        "Categoria":         df["Categoria"],
        "Valor (R$)":        df["Valor (R$)"],
        "Saldo (R$)":        df.get("Saldo (R$)", pd.Series([None]*len(df))),
    })
    df_final = df_final.drop_duplicates().reset_index(drop=True)
    return df_final


def gerar_excel_conciliacao(df):
    """Gera o Excel formatado da conciliação bancária."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_exp = df.copy()
        df_exp["Valor (R$)"] = df_exp["Valor (R$)"].apply(
            lambda x: f"R$ {float(x):,.2f}".replace(',','X').replace('.', ',').replace('X','.') if x else ""
        )
        df_exp.to_excel(writer, index=False, sheet_name="Conciliação")
        ws = writer.sheets["Conciliação"]
        verde = PatternFill(start_color="00C6EFCE", end_color="00C6EFCE", fill_type="solid")
        vermelho = PatternFill(start_color="00FFCCCC", end_color="00FFCCCC", fill_type="solid")
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="00203864", end_color="00203864", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        for row in ws.iter_rows(min_row=2):
            es = row[1].value if len(row) > 1 else ""
            fill = verde if es == "ENTRADA" else vermelho
            for cell in row:
                cell.fill = fill
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    buf.seek(0)
    return buf.read()


# ── DADOS MOCK ────────────────────────────────────────────────────────────────
MESES = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
RECEITA = [142000,158000,135000,172000,189000,201000,178000,215000,198000,234000,221000,267000]
DESPESA = [ 98000,112000, 95000,118000,125000,131000,119000,142000,128000,151000,139000,168000]
LUCRO   = [r - d for r, d in zip(RECEITA, DESPESA)]

TRANSACOES_MOCK = [
    {"id":"#001","data":"15/12/2025","descricao":"Venda — Cliente Alpha S.A.",   "categoria":"Receita", "valor":"R$ 45.200","variacao":"+18%","status":"Confirmado"},
    {"id":"#002","data":"14/12/2025","descricao":"Folha de Pagamento Dez/25",    "categoria":"Despesa", "valor":"R$ 32.500","variacao":"—",   "status":"Processado"},
    {"id":"#003","data":"13/12/2025","descricao":"Venda — Beta Tecnologia Ltda", "categoria":"Receita", "valor":"R$ 18.750","variacao":"+5%", "status":"Confirmado"},
    {"id":"#004","data":"12/12/2025","descricao":"Aluguel Escritório — Dez",     "categoria":"Despesa", "valor":"R$  8.900","variacao":"—",   "status":"Processado"},
    {"id":"#005","data":"11/12/2025","descricao":"Serviços — Gamma Corp",        "categoria":"Receita", "valor":"R$ 29.100","variacao":"+9%", "status":"Pendente"},
    {"id":"#006","data":"10/12/2025","descricao":"Fornecedor Insumos Ltda",      "categoria":"Despesa", "valor":"R$ 14.320","variacao":"—",   "status":"Processado"},
    {"id":"#007","data":"09/12/2025","descricao":"Venda — Delta Indústria",      "categoria":"Receita", "valor":"R$ 67.400","variacao":"+22%","status":"Confirmado"},
    {"id":"#008","data":"08/12/2025","descricao":"Marketing Digital — Nov",      "categoria":"Despesa", "valor":"R$  5.600","variacao":"—",   "status":"Processado"},
    {"id":"#009","data":"07/12/2025","descricao":"Venda — Epsilon Global",       "categoria":"Receita", "valor":"R$ 38.900","variacao":"+11%","status":"Confirmado"},
    {"id":"#010","data":"06/12/2025","descricao":"Manutenção Infraestrutura",    "categoria":"Despesa", "valor":"R$  9.200","variacao":"—",   "status":"Processado"},
]

DRE_MOCK = [
    {"conta":"RECEITA BRUTA","jan":"267.000","fev":"242.000","mar":"198.000","total":"707.000","tipo":"receita_bruta"},
]

CHART_LAYOUT = dict(
    paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
    font=dict(family='Sora, sans-serif', color='#8fa894', size=11),
    margin=dict(l=16, r=16, t=16, b=16), showlegend=True,
    legend=dict(bgcolor='rgba(0,0,0,0)', font=dict(color='#8fa894', size=10),
                orientation='h', yanchor='bottom', y=-0.3, xanchor='center', x=0.5),
    xaxis=dict(gridcolor='rgba(255,255,255,0.04)', zerolinecolor='rgba(255,255,255,0.06)',
               tickfont=dict(size=10), linecolor='rgba(255,255,255,0.06)'),
    yaxis=dict(gridcolor='rgba(255,255,255,0.04)', zerolinecolor='rgba(255,255,255,0.06)',
               tickfont=dict(size=10), linecolor='rgba(255,255,255,0.06)'),
    hoverlabel=dict(bgcolor='#1a1f1c', bordercolor='#2a3330',
                    font=dict(family='Space Mono', color='#e8ede9', size=11)),
)

def kpi_card(label, value, sub, badge_text, badge_cls, value_cls=""):
    return html.Div(className='kpi-card', children=[
        html.Div(className='kpi-top', children=[
            html.Div(label, className='kpi-label'),
            html.Div(badge_text, className=f'kpi-badge {badge_cls}'),
        ]),
        html.Div(value, className=f'kpi-value {value_cls}'),
        html.Div(sub, className='kpi-sub'),
        html.Div(className='kpi-bar'),
    ])

def chart_card(eyebrow, title, figure, height=300):
    return html.Div(style={
        'background':'#131714','border':'1px solid rgba(255,255,255,0.07)',
        'borderRadius':'10px','overflow':'hidden',
    }, children=[
        html.Div(style={
            'padding':'16px 22px 12px','borderBottom':'1px solid rgba(255,255,255,0.06)',
            'background':'#1a1f1c',
        }, children=[
            html.Div(eyebrow, style={
                'fontFamily':"'Space Mono',monospace",'fontSize':'9px',
                'color':'#3ddc84','textTransform':'uppercase','letterSpacing':'3px','marginBottom':'3px',
            }),
            html.Div(title, style={
                'fontFamily':"'Playfair Display',serif",'fontSize':'15px',
                'fontWeight':'700','color':'#e8ede9',
            }),
        ]),
        dcc.Graph(figure=figure, config={'displayModeBar':False}, style={'height':f'{height}px'}),
    ])

def stat_row(label, value, pct, up=True):
    color = '#3ddc84' if up else '#ff5c5c'
    arrow = '↑' if up else '↓'
    r,g,b = (61,220,132) if up else (255,92,92)
    return html.Div(style={
        'display':'flex','alignItems':'center','justifyContent':'space-between',
        'padding':'12px 0','borderBottom':'1px solid rgba(255,255,255,0.05)',
    }, children=[
        html.Div(label, style={'fontSize':'13px','color':'#8fa894'}),
        html.Div(style={'display':'flex','alignItems':'center','gap':'12px'}, children=[
            html.Div(value, style={'fontFamily':"'Space Mono',monospace",'fontSize':'12px','color':'#e8ede9'}),
            html.Div(f'{arrow} {pct}', style={
                'fontFamily':"'Space Mono',monospace",'fontSize':'10px','color':color,
                'background':f'rgba({r},{g},{b},0.1)','padding':'3px 8px','borderRadius':'4px',
            }),
        ]),
    ])

def gerar_template_csv():
    linhas = ["campo,valor"]
    for campo in DRE_CAMPOS.keys():
        linhas.append(f"{campo},0")
    return '\n'.join(linhas).encode('utf-8')

def gerar_excel(dre_data=None):
    if dre_data is None:
        dre_data = {}
    def v(key): return dre_data.get(key, 0.0) or 0.0
    def fmt(val): return f"R$ {val:,.2f}".replace(',','X').replace('.', ',').replace('X','.')
    def pct(val, total): return f"{(val/total*100):.2f}%" if total else "—"
    receita_merc   = v('venda_vista') + v('venda_prazo')
    outras_rec     = v('or_juros')+v('or_alug')+v('or_outras')+v('or_fundos')+v('or_bonif')
    receita_total  = receita_merc + outras_rec
    total_compras  = sum(v(k) for k in ['c_bebidas','c_bolos','c_carnes','c_cigarros','c_conv',
                         'c_estcoz','c_picoles','c_salgados','c_insumos','c_embal',
                         'c_recarga','c_bichos','c_buffet'])
    margem         = receita_total - total_compras
    total_cf       = sum(v(k) for k in ['cf_energia','cf_tel','cf_agua','cf_iptu','cf_sal',
                         'cf_enc','cf_imp','cf_seg','cf_diar'])
    total_cv       = sum(v(k) for k in ['cv_sist','cv_terc','cv_exped','cv_honor','cv_manut',
                         'cv_viag','cv_taxas','cv_unif','cv_limp','cv_alug','cv_caixa',
                         'cv_mora','cv_banco','cv_cart','cv_brindes','cv_utens','cv_veic',
                         'cv_segpred','cv_gerais','cv_gas','cv_estoque','cv_comiss'])
    total_custos   = total_cf + total_cv
    lucro_bruto    = margem - total_cf
    lucro_oper     = lucro_bruto - total_cv
    lucro_liq      = lucro_oper
    RT = receita_total if receita_total else 1
    DRE_ROWS = [
        ("1. RECEITA TOTAL", receita_total, pct(receita_total,RT), 'titulo'),
        ("2. RECEITA — VENDAS DE MERCADORIAS", receita_merc, pct(receita_merc,RT), 'secao'),
        ("   2.1  Vendas à Vista", v('venda_vista'), pct(v('venda_vista'),RT), 'item'),
        ("   2.2  Vendas a Prazo", v('venda_prazo'), pct(v('venda_prazo'),RT), 'item'),
        ("", None, None, 'spacer'),
        ("5. MARGEM DE CONTRIBUIÇÃO", margem, pct(margem,RT), 'resultado_green'),
        ("", None, None, 'spacer'),
        ("7.  RESULTADO LUCRO BRUTO EMPRESA", lucro_bruto, pct(lucro_bruto,RT), 'resultado_green'),
        ("9.  RESULTADO LÍQUIDO DA EMPRESA", lucro_liq, pct(lucro_liq,RT), 'resultado_green'),
    ]
    wb = openpyxl.Workbook()
    CG="FF3DDC84"; CD="FF0D0F0E"; CB2="FF131714"; CB3="FF1A1F1C"
    CT="FFE8EDE9"; CM="FF8FA894"; CB4="FF222720"; CR="FFFF5C5C"
    def xf(c): return PatternFill("solid", fgColor=c)
    def xb():
        s = Side(style='thin', color="FF2A3330")
        return Border(left=s, right=s, top=s, bottom=s)
    TIPO = {
        'titulo':          (CG,  CB2, True,  12),
        'secao':           (CG,  CB3, True,  11),
        'item':            (CT,  CB4, False, 10),
        'resultado_green': (CG,  CB2, True,  12),
        'spacer':          (CM,  CD,  False,  8),
    }
    ws = wb.active; ws.title = "DRE"; ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 46
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.merge_cells('B2:D3')
    ws['B2'] = "DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO"
    ws['B2'].font = Font(name='Calibri', bold=True, size=16, color=CG)
    ws['B2'].fill = xf(CD)
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 38
    for col, label in [('B','DESCRIÇÃO'), ('C','VALOR (R$)'), ('D','% RECEITA')]:
        c = ws[f'{col}6']
        c.value = label
        c.font = Font(name='Calibri', bold=True, size=10, color=CD)
        c.fill = xf(CG)
        c.alignment = Alignment(horizontal='left' if col=='B' else 'right', vertical='center')
        c.border = xb()
    ws.row_dimensions[6].height = 22
    for i, (label, valor, perc, tipo) in enumerate(DRE_ROWS):
        row = 7 + i
        txt_color, bg_color, bold, size = TIPO.get(tipo, (CT, CB4, False, 10))
        ws.row_dimensions[row].height = 8 if tipo == 'spacer' else 20
        c_desc = ws[f'B{row}']
        c_desc.value = label
        c_desc.font  = Font(name='Calibri', bold=bold, size=size, color=txt_color)
        c_desc.fill  = xf(bg_color)
        c_desc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c_desc.border = xb()
        c_val = ws[f'C{row}']
        if valor is not None:
            c_val.value = valor
            c_val.number_format = 'R$ #,##0.00'
        c_val.font  = Font(name='Calibri', bold=bold, size=size, color=txt_color)
        c_val.fill  = xf(bg_color)
        c_val.alignment = Alignment(horizontal='right', vertical='center')
        c_val.border = xb()
        c_pct = ws[f'D{row}']
        c_pct.value = perc or ''
        c_pct.font  = Font(name='Calibri', size=9, color=CM if tipo=='item' else txt_color)
        c_pct.fill  = xf(bg_color)
        c_pct.alignment = Alignment(horizontal='right', vertical='center')
        c_pct.border = xb()
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

def _input_style(width='260px'):
    return {
        'width': width, 'minWidth': width, 'padding': '9px 16px',
        'background': '#0d0f0e', 'border': '1px solid rgba(255,255,255,0.1)',
        'borderRadius': '6px', 'color': '#e8ede9',
        'fontFamily': "'Space Mono',monospace", 'fontSize': '13px',
        'textAlign': 'right', 'outline': 'none',
    }

def _campo(field_id, placeholder='0,00', width='260px'):
    return dcc.Input(id=field_id, type='text', placeholder=placeholder,
                     debounce=False, style=_input_style(width))

def _row_titulo(num, label, color='#3ddc84', bg='rgba(61,220,132,0.07)', size='14px'):
    return html.Tr(style={'background': bg, 'borderBottom': '1px solid rgba(255,255,255,0.08)'}, children=[
        html.Td(f'{num}. {label}', colSpan=2, style={
            'padding': '13px 20px', 'fontWeight': '700', 'fontSize': size,
            'color': color, 'fontFamily': "'Sora',sans-serif", 'letterSpacing': '0.3px',
        }),
    ])

def _row_sub(num, label, field_id):
    return html.Tr(style={'borderBottom': '1px solid rgba(255,255,255,0.04)'}, children=[
        html.Td(f'  {num}  {label}', style={
            'padding': '10px 20px', 'fontSize': '13px', 'width': '60%',
            'color': '#8fa894', 'fontFamily': "'Sora',sans-serif",
        }),
        html.Td(_campo(field_id), style={'padding': '6px 24px 6px 0', 'textAlign': 'right', 'width': '40%'}),
    ])

def _row_resultado(label, result_id, color='#e8c97a', bg='rgba(201,168,76,0.06)'):
    return html.Tr(style={'background': bg, 'borderTop': '2px solid rgba(255,255,255,0.1)', 'borderBottom': '1px solid rgba(255,255,255,0.08)'}, children=[
        html.Td(label, style={
            'padding': '13px 20px', 'fontWeight': '700', 'fontSize': '14px',
            'color': color, 'fontFamily': "'Sora',sans-serif",
        }),
        html.Td(id=result_id, children='—', style={
            'padding': '13px 20px', 'textAlign': 'right', 'fontWeight': '700',
            'fontSize': '14px', 'color': color, 'fontFamily': "'Space Mono',monospace",
        }),
    ])

def _build_dre_tab():
    # ── Painel de Indicadores Reais (toggleável) ──────────────────────────────
    indicadores_panel = html.Div(id='ind-panel-wrapper', children=[
        # Barra de controle
        html.Div(style={
            'display':'flex','alignItems':'center','justifyContent':'space-between',
            'marginBottom':'16px',
        }, children=[
            html.Div(style={'display':'flex','alignItems':'center','gap':'12px'}, children=[
                html.Div("INDICADORES DO DRE", style={
                    'fontFamily':"'Space Mono',monospace",'fontSize':'9px','color':'#3ddc84',
                    'textTransform':'uppercase','letterSpacing':'3px',
                }),
                html.Div("Calculados em tempo real a partir dos valores preenchidos", style={
                    'fontFamily':"'Space Mono',monospace",'fontSize':'9px','color':'#4d5e52',
                }),
            ]),
            html.Button(
                id='btn-toggle-ind', n_clicks=0,
                children="⊟ Ocultar Indicadores",
                style={
                    'padding':'7px 16px','borderRadius':'6px',
                    'border':'1px solid rgba(61,220,132,0.3)',
                    'background':'rgba(61,220,132,0.07)','color':'#3ddc84',
                    'fontFamily':"'Space Mono',monospace",'fontSize':'9px',
                    'fontWeight':'700','letterSpacing':'1px','cursor':'pointer',
                    'transition':'all 0.25s',
                }
            ),
        ]),

        # Cards de indicadores
        html.Div(id='ind-cards-wrapper', children=[
            # Linha 1 — 4 KPIs principais
            html.Div(className='kpi-grid', style={'marginBottom':'16px'}, children=[
                html.Div(id='ind-receita-total',   className='kpi-card'),
                html.Div(id='ind-margem-contrib',  className='kpi-card'),
                html.Div(id='ind-lucro-bruto',     className='kpi-card'),
                html.Div(id='ind-lucro-liquido',   className='kpi-card'),
            ]),
            # Linha 2 — 4 KPIs secundários
            html.Div(className='kpi-grid', style={'marginBottom':'24px'}, children=[
                html.Div(id='ind-total-compras',   className='kpi-card'),
                html.Div(id='ind-custos-fixos',    className='kpi-card'),
                html.Div(id='ind-custos-variaveis',className='kpi-card'),
                html.Div(id='ind-margem-pct',      className='kpi-card'),
            ]),
        ]),

        html.Div(className='divider'),
    ])

    upload_bar = dcc.Upload(
        id='upload-data',
        children=html.Div(style={'display':'flex','alignItems':'center','justifyContent':'space-between','width':'100%','gap':'16px'}, children=[
            html.Div(style={'display':'flex','alignItems':'center','gap':'12px'}, children=[
                html.Span("⬆", className='upload-icon'),
                html.Span(["Arraste ou clique para importar o ", html.Strong("DRE (.csv)"),
                           " — campos preenchidos automaticamente"], className='upload-text'),
            ]),
            html.Div(id='upload-status-text', children="AGUARDANDO CSV", className='upload-status'),
        ]),
        className='upload-area', multiple=False,
    )
    csv_info = html.Div(id='csv-banner', className='csv-info', children=[
        html.Div([html.Strong("Como importar via CSV: "), "Baixe o template em ", html.Strong("Baixar Template CSV"),
                  " na barra lateral, preencha os valores e importe aqui."]),
        html.Div([html.Strong("Formato: "), "duas colunas — ", html.Strong("campo"), " e ", html.Strong("valor"),
                  " (vírgula ou ponto-e-vírgula). Valores com R$, pontos e vírgulas são normalizados."]),
    ])
    return html.Div([
        indicadores_panel,
        upload_bar, csv_info,
        html.Div(className='section-header', children=[
            html.Div([
                html.Div("Demonstração do Resultado", className='section-eyebrow'),
                html.Div("DRE — Lançamento de Dados", className='section-title'),
            ]),
            html.Div("Preencha os valores e os resultados serão calculados automaticamente",
                     style={'fontSize':'12px','color':'#4d5e52','fontFamily':"'Space Mono',monospace"}),
        ]),
        html.Div(className='dre-card', children=[
            html.Div(className='dre-card-header', children=[
                html.Div(["DRE · ", html.Span("Entrada de Dados")], className='dre-card-title'),
                html.Div("Valores em R$", className='dre-pill'),
            ]),
            html.Table(style={'width':'100%','borderCollapse':'collapse'}, children=[html.Tbody([
                _row_titulo('1', 'RECEITA TOTAL', color='#3ddc84', bg='rgba(61,220,132,0.1)'),
                _row_titulo('2', 'RECEITA — VENDAS DE MERCADORIAS', bg='rgba(61,220,132,0.05)'),
                _row_sub('2.1', 'Vendas à Vista',  'dre-venda-vista'),
                _row_sub('2.2', 'Vendas a Prazo',  'dre-venda-prazo'),
                _row_titulo('3', 'VENDAS POR GRUPOS', bg='rgba(255,255,255,0.02)'),
                _row_sub('3.1',  'Bolos e Tortas',                'dre-g-bolos'),
                _row_sub('3.2',  'Buffet',                        'dre-g-buffet'),
                _row_sub('3.3',  'Cafés e Sucos',                 'dre-g-cafes'),
                _row_sub('3.4',  'Lanches',                       'dre-g-lanches'),
                _row_sub('3.5',  'Porções',                       'dre-g-porcoes'),
                _row_sub('3.6',  'Salgados Prontos',              'dre-g-salgados'),
                _row_sub('3.7',  'Drinks',                        'dre-g-drinks'),
                _row_sub('3.8',  'Conveniência',                  'dre-g-conv'),
                _row_sub('3.9',  'Cigarros',                      'dre-g-cigarros'),
                _row_sub('3.10', 'Bebidas',                       'dre-g-bebidas'),
                _row_sub('3.11', 'Picolé',                        'dre-g-picole'),
                _row_sub('3.12', 'Estoque Cozinha',               'dre-g-estcoz'),
                _row_sub('3.13', 'Bichos de Pelúcia e Brinquedos','dre-g-bichos'),
                _row_sub('3.14', 'Carnes',                        'dre-g-carnes'),
                _row_titulo('4', 'OUTRAS RECEITAS OPERACIONAIS', bg='rgba(255,255,255,0.02)'),
                _row_sub('4.1', 'Juros Recebidos',                'dre-or-juros'),
                _row_sub('4.2', 'Aluguéis',                       'dre-or-alug'),
                _row_sub('4.4', 'Outras Receitas',                'dre-or-outras'),
                _row_sub('4.5', 'Rendimentos Fundos de Investimentos', 'dre-or-fundos'),
                _row_sub('4.6', 'Bonificações Recebidas',         'dre-or-bonif'),
                _row_resultado('1. RECEITA TOTAL', 'res-receita-total'),
                _row_titulo('5', 'TOTAL COMPRA DE MERCADORIAS', color='#ff5c5c', bg='rgba(255,92,92,0.05)'),
                _row_sub('5.1',  'Bebidas',                       'dre-c-bebidas'),
                _row_sub('5.2',  'Bolos e Tortas',                'dre-c-bolos'),
                _row_sub('5.3',  'Carnes',                        'dre-c-carnes'),
                _row_sub('5.4',  'Cigarros',                      'dre-c-cigarros'),
                _row_sub('5.5',  'Conveniência',                  'dre-c-conv'),
                _row_sub('5.6',  'Estoque Cozinha',               'dre-c-estcoz'),
                _row_sub('5.7',  'Picolés',                       'dre-c-picoles'),
                _row_sub('5.8',  'Salgados Prontos',              'dre-c-salgados'),
                _row_sub('5.9',  'Insumos',                       'dre-c-insumos'),
                _row_sub('5.10', 'Embalagens',                    'dre-c-embal'),
                _row_sub('5.11', 'Recarga Celular',               'dre-c-recarga'),
                _row_sub('5.12', 'Bichos de Pelúcia e Brinquedos','dre-c-bichos'),
                _row_sub('5.13', 'Buffet',                        'dre-c-buffet'),
                _row_resultado('5. MARGEM DE CONTRIBUIÇÃO', 'res-margem', color='#3ddc84', bg='rgba(61,220,132,0.08)'),
                _row_titulo('6', 'TOTAL CUSTOS FIXOS + VARIÁVEIS', color='#ff5c5c', bg='rgba(255,92,92,0.05)'),
                _row_titulo('6.1', 'Custos Fixos', color='#ff8080', bg='rgba(255,92,92,0.03)', size='13px'),
                _row_sub('6.1.1', 'Energia Elétrica',             'dre-cf-energia'),
                _row_sub('6.1.2', 'Telefone',                     'dre-cf-tel'),
                _row_sub('6.1.3', 'Água',                         'dre-cf-agua'),
                _row_sub('6.1.4', 'IPTU',                         'dre-cf-iptu'),
                _row_sub('6.1.5', 'Salários / Férias / Rescisões / 13º', 'dre-cf-sal'),
                _row_sub('6.1.6', 'Encargos Sociais',             'dre-cf-enc'),
                _row_sub('6.1.7', 'Impostos',                     'dre-cf-imp'),
                _row_sub('6.1.8', 'Seguro de Vida — Funcionários','dre-cf-seg'),
                _row_sub('6.1.9', 'Diárias',                      'dre-cf-diar'),
                _row_titulo('6.2', 'Custos Variáveis', color='#ff8080', bg='rgba(255,92,92,0.03)', size='13px'),
                _row_sub('6.2.1',  'Despesas com Sistemas',        'dre-cv-sist'),
                _row_sub('6.2.2',  'Serviços de Terceiros',        'dre-cv-terc'),
                _row_sub('6.2.3',  'Materiais de Expediente',      'dre-cv-exped'),
                _row_sub('6.2.4',  'Honorários Contábeis',         'dre-cv-honor'),
                _row_sub('6.2.5',  'Manutenção e Reposição',       'dre-cv-manut'),
                _row_sub('6.2.6',  'Viagens e Estadias',           'dre-cv-viag'),
                _row_sub('6.2.7',  'Taxas Diversas',               'dre-cv-taxas'),
                _row_sub('6.2.8',  'Uniformes',                    'dre-cv-unif'),
                _row_sub('6.2.9',  'Material de Limpeza',          'dre-cv-limp'),
                _row_sub('6.2.10', 'Aluguéis e Locações',          'dre-cv-alug'),
                _row_sub('6.2.11', 'Faltas e Sobras de Caixa',     'dre-cv-caixa'),
                _row_sub('6.2.12', 'Juros de Mora',                'dre-cv-mora'),
                _row_sub('6.2.13', 'Tarifas Bancárias',            'dre-cv-banco'),
                _row_sub('6.2.14', 'Tarifas Cartões',              'dre-cv-cart'),
                _row_sub('6.2.15', 'Brindes e Bonificações',       'dre-cv-brindes'),
                _row_sub('6.2.16', 'Utensílios',                   'dre-cv-utens'),
                _row_sub('6.2.17', 'Despesas com Veículos',        'dre-cv-veic'),
                _row_sub('6.2.18', 'Seguros Edificações',          'dre-cv-segpred'),
                _row_sub('6.2.19', 'Despesas Gerais',              'dre-cv-gerais'),
                _row_sub('6.2.20', 'Gás GLP',                      'dre-cv-gas'),
                _row_sub('6.2.21', 'Contagem Estoque',             'dre-cv-estoque'),
                _row_sub('6.2.22', 'Comissões sobre Vendas',       'dre-cv-comiss'),
                _row_resultado('7. LUCRO BRUTO DA EMPRESA',   'res-lucro-bruto', color='#3ddc84', bg='rgba(61,220,132,0.08)'),
                _row_resultado('8. RESULTADO OPERACIONAL',    'res-operacional', color='#3ddc84', bg='rgba(61,220,132,0.06)'),
                _row_resultado('9. RESULTADO LÍQUIDO',        'res-liquido',     color='#3ddc84', bg='rgba(61,220,132,0.12)'),
                _row_titulo('10', 'FUNCIONÁRIOS & HORAS EXTRAS', color='#5c9eff', bg='rgba(92,158,255,0.06)'),
                html.Tr(style={'borderBottom':'1px solid rgba(255,255,255,0.04)'}, children=[
                    html.Td('  10.  Total de Funcionários', style={'padding':'10px 20px','fontSize':'13px','color':'#8fa894','fontFamily':"'Sora',sans-serif"}),
                    html.Td(_campo('dre-func-total', placeholder='Nº funcionários', width='160px'), style={'padding':'6px 20px 6px 0','textAlign':'right'}),
                ]),
                html.Tr(style={'borderBottom':'1px solid rgba(255,255,255,0.04)'}, children=[
                    html.Td('  11.  Total de Horas Extras', style={'padding':'10px 20px','fontSize':'13px','color':'#8fa894','fontFamily':"'Sora',sans-serif"}),
                    html.Td(_campo('dre-he-qtd', placeholder='Total de horas', width='160px'), style={'padding':'6px 20px 6px 0','textAlign':'right'}),
                ]),
                html.Tr(style={'background':'rgba(92,158,255,0.05)','borderBottom':'2px solid rgba(92,158,255,0.15)'}, children=[
                    html.Td('  12.  Valor Total de Horas Extras', style={'padding':'12px 20px','fontSize':'13px','fontWeight':'600','color':'#5c9eff','fontFamily':"'Sora',sans-serif"}),
                    html.Td(style={'padding':'6px 20px 6px 0','textAlign':'right','display':'flex','gap':'8px','justifyContent':'flex-end'}, children=[
                        _campo('dre-he-total', placeholder='R$ Valor total', width='160px'),
                        dcc.Input(id='dre-he-pct', type='number', placeholder='%', debounce=True,
                            style={'width':'80px','padding':'7px 10px','background':'#0d0f0e',
                                'border':'1px solid rgba(92,158,255,0.3)','borderRadius':'6px',
                                'color':'#5c9eff','fontFamily':"'Space Mono',monospace",
                                'fontSize':'12px','textAlign':'right','outline':'none'}),
                    ]),
                ]),
            ])]),
        ]),
    ])


# ══════════════════════════════════════════════════════════════════════════════
# ABA CONCILIAÇÃO BANCÁRIA
# ══════════════════════════════════════════════════════════════════════════════

TIPOS_DISPONIVEIS = [
    "TODOS", "PIX", "TED", "DOC", "CARTÃO", "BOLETOS RECEBIDOS", "BOLETOS PAGOS",
    "DEPÓSITO", "PAGSEGURO", "ESTORNO", "TARIFA BANCÁRIA", "TARIFA BOLETO",
    "DEB. AUTOMÁTICO", "IOF", "JUROS/MULTA", "OUTROS VALORES", "OUTROS",
]

TYPE_COLORS = {
    "PIX":               "#3ddc84",
    "TED":               "#5c9eff",
    "DOC":               "#5c9eff",
    "CARTÃO":            "#c9a84c",
    "BOLETOS RECEBIDOS": "#3ddc84",
    "BOLETOS PAGOS":     "#ff5c5c",
    "DEPÓSITO":          "#3ddc84",
    "PAGSEGURO":         "#3ddc84",
    "ESTORNO":           "#ff8080",
    "TARIFA BANCÁRIA":   "#ff5c5c",
    "TARIFA BOLETO":     "#ff5c5c",
    "DEB. AUTOMÁTICO":   "#ff5c5c",
    "IOF":               "#ff5c5c",
    "JUROS/MULTA":       "#ff5c5c",
    "OUTROS VALORES":    "#8fa894",
    "OUTROS":            "#8fa894",
}

def _build_conciliacao_tab():
    dropdown_style = {
        'background':'#0d0f0e','color':'#e8ede9',
        'border':'1px solid rgba(255,255,255,0.1)','borderRadius':'6px',
        'fontFamily':"'Space Mono',monospace",'fontSize':'11px',
    }
    return html.Div([
        # Upload
        dcc.Upload(
            id='upload-extrato',
            children=html.Div(style={'display':'flex','alignItems':'center','justifyContent':'space-between','width':'100%','gap':'16px'}, children=[
                html.Div(style={'display':'flex','alignItems':'center','gap':'12px'}, children=[
                    html.Span("🏦", style={'fontSize':'20px'}),
                    html.Div([
                        html.Div(["Arraste ou clique para importar o ", html.Strong("Extrato Bancário (.xlsx / .xls)")],
                                 style={'fontSize':'13px','color':'#8fa894'}),
                        html.Div("Colunas esperadas: Data · Descrição · Documento · Valor (R$) · Saldo (R$) — a partir da linha 9",
                                 style={'fontSize':'10px','color':'#4d5e52','marginTop':'3px','fontFamily':"'Space Mono',monospace"}),
                    ]),
                ]),
                html.Div(id='extrato-status-text', children="AGUARDANDO EXTRATO", className='upload-status'),
            ]),
            className='upload-area', multiple=False,
            style={'marginBottom':'20px'},
        ),

        # KPIs resumo
        html.Div(id='conc-kpis', style={'marginBottom':'20px'}),

        # Filtros
        html.Div(style={
            'display':'flex','gap':'12px','alignItems':'flex-end','marginBottom':'20px',
            'flexWrap':'wrap',
        }, children=[
            html.Div([
                html.Div("Filtrar por Tipo", style={'fontFamily':"'Space Mono',monospace",'fontSize':'9px','color':'#4d5e52','textTransform':'uppercase','letterSpacing':'2px','marginBottom':'6px'}),
                dcc.Dropdown(
                    id='filtro-tipo',
                    options=[{'label': t, 'value': t} for t in TIPOS_DISPONIVEIS],
                    value='TODOS',
                    clearable=False,
                    style={'width':'220px','background':'#0d0f0e'},
                    className='dre-dropdown',
                ),
            ]),
            html.Div([
                html.Div("Filtrar por Entrada/Saída", style={'fontFamily':"'Space Mono',monospace",'fontSize':'9px','color':'#4d5e52','textTransform':'uppercase','letterSpacing':'2px','marginBottom':'6px'}),
                dcc.Dropdown(
                    id='filtro-es',
                    options=[
                        {'label': 'TODOS', 'value': 'TODOS'},
                        {'label': '⬆ ENTRADA', 'value': 'ENTRADA'},
                        {'label': '⬇ SAÍDA',   'value': 'SAÍDA'},
                    ],
                    value='TODOS',
                    clearable=False,
                    style={'width':'180px'},
                    className='dre-dropdown',
                ),
            ]),
            html.Div([
                html.Div("Buscar na Descrição", style={'fontFamily':"'Space Mono',monospace",'fontSize':'9px','color':'#4d5e52','textTransform':'uppercase','letterSpacing':'2px','marginBottom':'6px'}),
                dcc.Input(
                    id='filtro-busca',
                    type='text',
                    placeholder='Filtrar descrição...',
                    debounce=True,
                    style={
                        'padding':'9px 14px','background':'#0d0f0e',
                        'border':'1px solid rgba(255,255,255,0.1)','borderRadius':'6px',
                        'color':'#e8ede9','fontFamily':"'Space Mono',monospace",'fontSize':'11px',
                        'outline':'none','width':'220px',
                    }
                ),
            ]),
            # Botão exportar
            html.Div(style={'marginLeft':'auto','display':'flex','alignItems':'flex-end'}, children=[
                html.Button(
                    "⬇ Exportar Excel",
                    id='btn-export-conc',
                    className='btn-action',
                    style={'padding':'9px 18px','fontSize':'11px'},
                ),
            ]),
        ]),

        # Gráficos
        html.Div(id='conc-charts', style={'marginBottom':'20px'}),

        # Tabela
        html.Div(id='conc-table'),

        # Download
        dcc.Download(id='download-conc-excel'),
    ])


# ── INDEX STRING ──────────────────────────────────────────────────────────────
app.index_string = '''<!DOCTYPE html>
<html>
    <head>
        {%metas%}
        <title>GUAPO · ERPWeb</title>
        {%favicon%}
        {%css%}
        <link rel="preconnect" href="https://fonts.googleapis.com">
        <link href="https://fonts.googleapis.com/css2?family=Playfair+Display:ital,wght@0,400;0,700;0,900;1,400&family=Space+Mono:ital,wght@0,400;0,700;1,400&family=Sora:wght@300;400;500;600;700&display=swap" rel="stylesheet">
        <style>
            *,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
            :root{
                --bg:#0d0f0e;--bg2:#131714;--bg3:#1a1f1c;--bg4:#222720;
                --border:rgba(255,255,255,0.07);--border-lit:rgba(61,220,132,0.25);
                --green:#3ddc84;--green-dim:#2ab86a;--green-glow:rgba(61,220,132,0.15);
                --green-deep:rgba(61,220,132,0.06);--gold:#c9a84c;--gold-light:#e8c97a;
                --text:#e8ede9;--text-mid:#8fa894;--text-dim:#4d5e52;
                --red:#ff5c5c;--blue:#5c9eff;
                --shadow:0 0 0 1px rgba(255,255,255,0.04),0 8px 32px rgba(0,0,0,0.4);
                --shadow-lg:0 0 0 1px rgba(255,255,255,0.05),0 24px 64px rgba(0,0,0,0.6);
                --r:10px;--t:0.3s cubic-bezier(0.16,1,0.3,1);
            }
            html{scroll-behavior:smooth}
            body{font-family:'Sora',sans-serif;background:var(--bg);color:var(--text);-webkit-font-smoothing:antialiased;min-height:100vh}
            ::-webkit-scrollbar{width:5px}::-webkit-scrollbar-track{background:var(--bg2)}
            ::-webkit-scrollbar-thumb{background:var(--bg4);border-radius:3px}
            ::-webkit-scrollbar-thumb:hover{background:var(--green-dim)}

            .screen-login{min-height:100vh;display:flex;flex-direction:column;position:relative;overflow:hidden;background:var(--bg)}
            .screen-login::before{content:'';position:absolute;inset:0;background-image:linear-gradient(rgba(61,220,132,0.03) 1px,transparent 1px),linear-gradient(90deg,rgba(61,220,132,0.03) 1px,transparent 1px);background-size:48px 48px;animation:grid-drift 20s linear infinite;pointer-events:none}
            @keyframes grid-drift{0%{background-position:0 0}100%{background-position:48px 48px}}
            .screen-login::after{content:'';position:absolute;top:-30vh;right:-20vw;width:70vw;height:70vw;border-radius:50%;background:radial-gradient(circle,rgba(61,220,132,0.06) 0%,transparent 65%);pointer-events:none}
            .login-top-bar{position:relative;z-index:2;padding:18px 48px;display:flex;align-items:center;justify-content:space-between;border-bottom:1px solid var(--border)}
            .login-top-brand{font-family:'Playfair Display',serif;font-size:20px;font-weight:700;color:var(--green)}
            .login-top-tag{font-family:'Space Mono',monospace;font-size:10px;color:var(--text-dim);text-transform:uppercase;letter-spacing:3px}
            .login-top-dot{width:8px;height:8px;border-radius:50%;background:var(--green);box-shadow:0 0 12px var(--green);animation:blink 2s ease-in-out infinite}
            @keyframes blink{0%,100%{opacity:1}50%{opacity:0.3}}
            .login-center{flex:1;display:flex;align-items:center;justify-content:center;position:relative;z-index:2;padding:40px 20px}
            .login-panel{width:440px;animation:panel-in 0.7s cubic-bezier(0.16,1,0.3,1) both}
            @keyframes panel-in{from{opacity:0;transform:translateY(32px) scale(0.97)}to{opacity:1;transform:translateY(0) scale(1)}}
            .login-eyebrow{font-family:'Space Mono',monospace;font-size:10px;color:var(--green);text-transform:uppercase;letter-spacing:4px;margin-bottom:16px;display:flex;align-items:center;gap:10px}
            .login-eyebrow::before{content:'';display:block;width:28px;height:1px;background:var(--green);opacity:0.5}
            .login-headline{font-family:'Playfair Display',serif;font-size:52px;font-weight:900;line-height:1;letter-spacing:-2px;margin-bottom:8px;color:var(--text)}
            .login-headline em{font-style:italic;color:var(--green)}
            .login-desc{font-size:13px;color:var(--text-mid);font-weight:300;margin-bottom:40px;letter-spacing:0.3px}
            .login-card{background:var(--bg2);border:1px solid var(--border);border-radius:16px;padding:36px;box-shadow:var(--shadow-lg);position:relative;overflow:hidden}
            .login-card::before{content:'';position:absolute;top:0;left:0;right:0;height:1px;background:linear-gradient(90deg,transparent,var(--green),transparent);opacity:0.4}
            .field-group{margin-bottom:20px}
            .field-label{font-family:'Space Mono',monospace;font-size:10px;color:var(--text-dim);text-transform:uppercase;letter-spacing:2px;margin-bottom:8px;display:block}
            .input-field{width:100%!important;padding:14px 18px!important;border-radius:8px!important;border:1px solid var(--border)!important;font-family:'Sora',sans-serif!important;font-size:14px!important;color:var(--text)!important;background:var(--bg3)!important;outline:none!important;display:block!important;transition:border-color var(--t),box-shadow var(--t),background var(--t)!important}
            .input-field:focus{border-color:var(--green-dim)!important;background:var(--bg4)!important;box-shadow:0 0 0 3px var(--green-deep),0 0 20px rgba(61,220,132,0.05)!important}
            .input-field::placeholder{color:var(--text-dim)!important}
            .btn-login{width:100%;margin-top:8px;padding:16px 24px;background:var(--green);color:var(--bg);border:none;border-radius:8px;cursor:pointer;font-family:'Space Mono',monospace;font-weight:700;font-size:12px;letter-spacing:3px;text-transform:uppercase;transition:all var(--t)}
            .btn-login:hover{background:var(--gold-light);transform:translateY(-1px);box-shadow:0 8px 24px rgba(61,220,132,0.3)}
            .login-error{margin-top:16px;min-height:20px;font-family:'Space Mono',monospace;font-size:11px;color:var(--red);text-align:center;letter-spacing:1px}
            .login-footer-note{margin-top:20px;text-align:center;font-family:'Space Mono',monospace;font-size:10px;color:var(--text-dim);letter-spacing:1px}

            .screen-dash{background:var(--bg);min-height:100vh}
            .dash-layout{display:flex}
            .sidebar{width:240px;flex-shrink:0;background:var(--bg2);border-right:1px solid var(--border);display:flex;flex-direction:column;position:fixed;top:0;left:0;bottom:0;z-index:50;overflow-y:auto}
            .sidebar-brand{padding:28px 24px 22px;border-bottom:1px solid var(--border)}
            .sidebar-logo{font-family:'Playfair Display',serif;font-size:30px;font-weight:900;color:var(--text);letter-spacing:-1px;line-height:1}
            .sidebar-logo span{color:var(--green)}
            .sidebar-sub{font-family:'Space Mono',monospace;font-size:9px;color:var(--text-dim);text-transform:uppercase;letter-spacing:3px;margin-top:5px}
            .sidebar-section{padding:22px 14px 8px}
            .sidebar-section-label{font-family:'Space Mono',monospace;font-size:9px;color:var(--text-dim);text-transform:uppercase;letter-spacing:3px;padding:0 10px;margin-bottom:8px}
            .nav-item{display:flex;align-items:center;gap:12px;padding:11px 12px;border-radius:8px;cursor:pointer;margin-bottom:2px;transition:background var(--t),color var(--t);font-size:13px;font-weight:500;color:var(--text-mid);border:1px solid transparent;user-select:none}
            .nav-item:hover{background:var(--bg3);color:var(--text)}
            .nav-item.active{background:var(--green-deep);border-color:var(--border-lit);color:var(--green)}
            .nav-icon{font-size:15px;width:20px;text-align:center}
            .sidebar-bottom{margin-top:auto;padding:18px 14px;border-top:1px solid var(--border)}
            .sidebar-user{display:flex;align-items:center;gap:12px;padding:10px 12px;border-radius:8px;background:var(--bg3)}
            .user-avatar{width:34px;height:34px;border-radius:50%;background:linear-gradient(135deg,var(--green-dim),var(--gold));display:flex;align-items:center;justify-content:center;font-size:14px;font-weight:700;color:var(--bg);flex-shrink:0}
            .user-name{font-size:13px;font-weight:600;color:var(--text)}
            .user-role{font-family:'Space Mono',monospace;font-size:9px;color:var(--text-dim);text-transform:uppercase;letter-spacing:1px}
            .dash-main{margin-left:240px;flex:1;min-height:100vh}
            .topbar{height:64px;flex-shrink:0;border-bottom:1px solid var(--border);padding:0 40px;display:flex;align-items:center;justify-content:space-between;background:rgba(13,15,14,0.9);backdrop-filter:blur(16px);position:sticky;top:0;z-index:40}
            .topbar-title{font-size:15px;font-weight:600;color:var(--text)}
            .topbar-subtitle{font-family:'Space Mono',monospace;font-size:10px;color:var(--text-dim);letter-spacing:1px;margin-top:2px}
            .topbar-actions{display:flex;align-items:center;gap:12px}
            .status-chip{display:flex;align-items:center;gap:8px;padding:7px 14px;border-radius:100px;background:var(--green-deep);border:1px solid var(--border-lit);font-family:'Space Mono',monospace;font-size:10px;color:var(--green);text-transform:uppercase;letter-spacing:2px}
            .status-dot{width:6px;height:6px;border-radius:50%;background:var(--green);box-shadow:0 0 8px var(--green);animation:blink 2s ease-in-out infinite}
            .topbar-date{font-family:'Space Mono',monospace;font-size:10px;color:var(--text-dim);letter-spacing:1px}
            .main-content{padding:36px 40px;flex:1}
            .upload-area{border:1px dashed rgba(61,220,132,0.2);border-radius:var(--r);background:var(--green-deep);padding:18px 24px;cursor:pointer;margin-bottom:32px;transition:border-color var(--t),background var(--t);display:flex;align-items:center;justify-content:space-between;gap:12px}
            .upload-area:hover{border-color:rgba(61,220,132,0.4);background:rgba(61,220,132,0.08)}
            .upload-icon{font-size:18px}
            .upload-text{font-size:13px;color:var(--text-mid);font-weight:400}
            .upload-text strong{color:var(--green);font-weight:600}
            .upload-status{font-family:'Space Mono',monospace;font-size:10px;color:var(--text-dim);letter-spacing:1px;white-space:nowrap}
            .upload-status.loaded{color:var(--green)}
            .section-header{display:flex;align-items:flex-end;justify-content:space-between;margin-bottom:24px}
            .section-eyebrow{font-family:'Space Mono',monospace;font-size:9px;color:var(--green);text-transform:uppercase;letter-spacing:3px;margin-bottom:6px}
            .section-title{font-family:'Playfair Display',serif;font-size:24px;font-weight:700;color:var(--text);letter-spacing:-0.5px}
            .section-period{font-family:'Space Mono',monospace;font-size:10px;color:var(--text-dim);letter-spacing:1px}
            .kpi-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:24px}
            .kpi-grid-3{display:grid;grid-template-columns:repeat(3,1fr);gap:16px;margin-bottom:24px}
            .kpi-grid-5{display:grid;grid-template-columns:repeat(5,1fr);gap:16px;margin-bottom:24px}
            .kpi-card{background:var(--bg2);border:1px solid var(--border);border-radius:var(--r);padding:24px 22px;position:relative;overflow:hidden;transition:border-color var(--t),transform var(--t),box-shadow var(--t);cursor:default}
            .kpi-card:hover{border-color:var(--border-lit);transform:translateY(-2px);box-shadow:0 12px 32px rgba(0,0,0,0.3)}
            .kpi-card::after{content:'';position:absolute;top:0;right:0;width:48px;height:48px;background:radial-gradient(circle at top right,rgba(61,220,132,0.08),transparent 70%);pointer-events:none}
            .kpi-top{display:flex;align-items:center;justify-content:space-between;margin-bottom:14px}
            .kpi-label{font-family:'Space Mono',monospace;font-size:9px;color:var(--text-dim);text-transform:uppercase;letter-spacing:2px}
            .kpi-badge{font-family:'Space Mono',monospace;font-size:9px;padding:3px 8px;border-radius:4px;font-weight:700;text-transform:uppercase;letter-spacing:1px}
            .kpi-badge.up{background:rgba(61,220,132,0.12);color:var(--green)}
            .kpi-badge.wait{background:rgba(201,168,76,0.12);color:var(--gold)}
            .kpi-badge.down{background:rgba(255,92,92,0.12);color:var(--red)}
            .kpi-badge.blue{background:rgba(92,158,255,0.12);color:var(--blue)}
            .kpi-value{font-family:'Playfair Display',serif;font-size:28px;font-weight:700;color:var(--text);line-height:1;letter-spacing:-0.5px}
            .kpi-value.green{color:var(--green)}
            .kpi-value.gold{color:var(--gold-light)}
            .kpi-value.red{color:var(--red)}
            .kpi-value.blue{color:var(--blue)}
            .kpi-sub{font-family:'Space Mono',monospace;font-size:10px;color:var(--text-dim);margin-top:8px;letter-spacing:0.5px}
            .kpi-bar{position:absolute;bottom:0;left:0;right:0;height:2px;background:linear-gradient(90deg,var(--green),transparent);opacity:0;transition:opacity var(--t)}
            .kpi-card:hover .kpi-bar{opacity:1}
            .divider{height:1px;background:var(--border);margin:28px 0}
            .charts-grid-2{display:grid;grid-template-columns:repeat(2,1fr);gap:20px;margin-bottom:24px}
            .dre-card{background:var(--bg2);border:1px solid var(--border);border-radius:var(--r);overflow:hidden}
            .dre-card-header{padding:18px 24px;border-bottom:1px solid var(--border);display:flex;align-items:center;justify-content:space-between;background:var(--bg3)}
            .dre-card-title{font-family:'Space Mono',monospace;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:2px;color:var(--text-mid)}
            .dre-card-title span{color:var(--green)}
            .dre-pill{font-family:'Space Mono',monospace;font-size:9px;color:var(--gold);background:rgba(201,168,76,0.1);border:1px solid rgba(201,168,76,0.2);padding:4px 12px;border-radius:4px;text-transform:uppercase;letter-spacing:1px}
            .table-card{background:var(--bg2);border:1px solid var(--border);border-radius:var(--r);overflow:hidden}
            .table-header{padding:18px 24px;border-bottom:1px solid var(--border);background:var(--bg3);display:flex;align-items:center;justify-content:space-between}
            .table-header-title{font-family:'Space Mono',monospace;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:2px;color:var(--text-mid)}
            .btn-action{padding:8px 18px;border-radius:6px;border:1px solid var(--border-lit);background:var(--green-deep);color:var(--green);font-family:'Space Mono',monospace;font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:2px;cursor:pointer;transition:all var(--t)}
            .btn-action:hover{background:var(--green-glow);border-color:var(--green);transform:translateY(-1px)}
            .btn-danger{padding:8px 18px;border-radius:6px;border:1px solid rgba(255,92,92,0.3);background:rgba(255,92,92,0.06);color:var(--red);font-family:'Space Mono',monospace;font-size:9px;font-weight:700;text-transform:uppercase;letter-spacing:2px;cursor:pointer;transition:all var(--t)}
            .btn-danger:hover{background:rgba(255,92,92,0.12);transform:translateY(-1px)}
            .toast{position:fixed;bottom:28px;right:28px;background:var(--bg3);border:1px solid var(--border-lit);border-radius:10px;padding:14px 20px;display:flex;align-items:center;gap:12px;font-size:13px;color:var(--text);box-shadow:var(--shadow-lg);z-index:9999;animation:toast-in 0.4s cubic-bezier(0.16,1,0.3,1)}
            @keyframes toast-in{from{opacity:0;transform:translateY(20px)}to{opacity:1;transform:translateY(0)}}
            .modal-overlay{position:fixed;inset:0;background:rgba(0,0,0,0.75);backdrop-filter:blur(8px);z-index:200;display:flex;align-items:center;justify-content:center;animation:fade-in 0.2s ease}
            @keyframes fade-in{from{opacity:0}to{opacity:1}}
            .modal-box{background:var(--bg2);border:1px solid var(--border);border-radius:16px;padding:36px;width:500px;box-shadow:var(--shadow-lg);position:relative;animation:modal-in 0.35s cubic-bezier(0.16,1,0.3,1)}
            @keyframes modal-in{from{opacity:0;transform:scale(0.94)}to{opacity:1;transform:scale(1)}}
            .modal-title{font-family:'Playfair Display',serif;font-size:22px;font-weight:700;color:var(--text);margin-bottom:6px}
            .modal-sub{font-size:13px;color:var(--text-mid);margin-bottom:24px;padding-bottom:20px;border-bottom:1px solid var(--border)}
            .modal-close{position:absolute;top:18px;right:22px;background:none;border:none;color:var(--text-dim);cursor:pointer;font-size:20px;line-height:1;transition:color var(--t)}
            .modal-close:hover{color:var(--text)}
            .config-row{display:flex;align-items:center;justify-content:space-between;padding:14px 0;border-bottom:1px solid var(--border)}
            .config-row:last-child{border-bottom:none;padding-bottom:0}
            .config-label{font-size:13px;font-weight:500;color:var(--text)}
            .config-sub{font-size:11px;color:var(--text-dim);margin-top:2px}
            .toggle{width:42px;height:24px;border-radius:12px;border:none;cursor:pointer;position:relative;transition:background var(--t);flex-shrink:0}
            .toggle.on{background:var(--green)}
            .toggle.off{background:var(--bg4);border:1px solid var(--border)}
            .toggle::after{content:'';position:absolute;top:4px;width:16px;height:16px;border-radius:50%;background:white;transition:left 0.2s}
            .toggle.on::after{left:22px}
            .toggle.off::after{left:4px}
            tr:hover td{background:rgba(255,255,255,0.02)!important}
            .csv-info{background:rgba(61,220,132,0.04);border:1px solid rgba(61,220,132,0.12);border-radius:8px;padding:12px 16px;margin-bottom:16px;font-family:'Space Mono',monospace;font-size:10px;color:var(--text-dim);line-height:1.7}
            .csv-info strong{color:var(--green)}
            /* Conciliação table */
            .conc-table-wrap{overflow-x:auto;border-radius:var(--r);border:1px solid var(--border)}
            .conc-table{width:100%;border-collapse:collapse;font-size:12px}
            .conc-table th{background:var(--bg3);padding:10px 14px;font-family:'Space Mono',monospace;font-size:9px;color:var(--text-dim);text-transform:uppercase;letter-spacing:2px;text-align:left;border-bottom:2px solid rgba(255,255,255,0.06);white-space:nowrap}
            .conc-table td{padding:10px 14px;border-bottom:1px solid rgba(255,255,255,0.04);color:var(--text-mid);vertical-align:middle;white-space:nowrap}
            .conc-table tr:hover td{background:rgba(255,255,255,0.025)!important}
            .tipo-badge{display:inline-block;padding:3px 9px;border-radius:4px;font-family:'Space Mono',monospace;font-size:9px;font-weight:700;letter-spacing:1px}
            .es-entrada{color:#3ddc84;font-weight:700;font-family:'Space Mono',monospace;font-size:10px}
            .es-saida{color:#ff5c5c;font-weight:700;font-family:'Space Mono',monospace;font-size:10px}
            .val-entrada{color:#3ddc84;font-family:'Space Mono',monospace;font-size:11px;font-weight:700;text-align:right}
            .val-saida{color:#ff5c5c;font-family:'Space Mono',monospace;font-size:11px;font-weight:700;text-align:right}
            .val-saldo{color:var(--text);font-family:'Space Mono',monospace;font-size:11px;text-align:right}

            /* Dropdown dark */
            .dre-dropdown .Select-control{background:#0d0f0e!important;border-color:rgba(255,255,255,0.1)!important;color:#e8ede9!important}
            .dre-dropdown .Select-menu-outer{background:#131714!important;border-color:rgba(255,255,255,0.1)!important}
            .dre-dropdown .Select-option{color:#8fa894!important;background:#131714!important}
            .dre-dropdown .Select-option.is-focused{background:#1a1f1c!important;color:#e8ede9!important}
            .dre-dropdown .Select-value-label{color:#e8ede9!important}
        </style>
        <script>
        document.addEventListener('DOMContentLoaded', function() {
            function attachBRL() {
                document.querySelectorAll('.dre-card input[type=text]').forEach(function(inp) {
                    if (inp._brlFmt) return;
                    inp._brlFmt = true;
                    inp.addEventListener('focus', function() {
                        var raw = this.value.replace(/[.]/g, '').replace(',', '.');
                        var num = parseFloat(raw);
                        this.value = (!isNaN(num) && num !== 0) ? num.toFixed(2).replace('.', ',') : '';
                    });
                    inp.addEventListener('blur', function() {
                        var raw = this.value.replace(/[.]/g, '').replace(',', '.');
                        var num = parseFloat(raw);
                        if (!isNaN(num)) {
                            this.value = num.toLocaleString('pt-BR', {minimumFractionDigits:2, maximumFractionDigits:2});
                        }
                    });
                });
            }
            attachBRL();
            new MutationObserver(attachBRL).observe(document.body, {childList:true, subtree:true});
        });
        </script>
    </head>
    <body>
        {%app_entry%}
        <footer>{%config%}{%scripts%}{%renderer%}</footer>
    </body>
</html>'''

# ── LAYOUT ─────────────────────────────────────────────────────────────────────
app.layout = html.Div([
    dcc.Store(id='auth-status', data=False),
    dcc.Store(id='csv-store',   data={}),
    dcc.Store(id='dre-store',   data={}),
    dcc.Store(id='active-tab',  data='tab-home'),
    dcc.Store(id='dre-log',     data={}),
    dcc.Store(id='extrato-store', data=None),   # ← dados da conciliação
    dcc.Interval(id='clock-tick', interval=60_000, n_intervals=0),
    dcc.Download(id='download-excel'),
    dcc.Download(id='download-template'),

    html.Div(id='toast', style={'display':'none'}),

    # MODAL CONFIGURAÇÕES
    html.Div(id='modal-config-overlay', style={'display':'none'}, children=[
        html.Div(className='modal-overlay', children=[
            html.Div(className='modal-box', children=[
                html.Button("✕", id='btn-modal-config-close', className='modal-close'),
                html.Div("Configurações", className='modal-title'),
                html.Div("Preferências do sistema GUAPO ERP", className='modal-sub'),
                html.Div([
                    html.Div(className='config-row', children=[
                        html.Div([html.Div("Notificações em Tempo Real", className='config-label'),
                                  html.Div("Alertas de variação e alertas críticos", className='config-sub')]),
                        html.Button(id='toggle-notif', className='toggle on'),
                    ]),
                    html.Div(className='config-row', children=[
                        html.Div([html.Div("Tema Escuro", className='config-label'),
                                  html.Div("Interface em modo dark (padrão)", className='config-sub')]),
                        html.Button(id='toggle-dark', className='toggle on'),
                    ]),
                    html.Div(className='config-row', children=[
                        html.Div([html.Div("Exportação Automática", className='config-label'),
                                  html.Div("Gerar Excel ao importar CSV", className='config-sub')]),
                        html.Button(id='toggle-autoexp', className='toggle off'),
                    ]),
                    html.Div(className='config-row', children=[
                        html.Div([html.Div("Atualização Automática", className='config-label'),
                                  html.Div("Recarregar dados a cada 5 minutos", className='config-sub')]),
                        html.Button(id='toggle-auto', className='toggle off'),
                    ]),
                ]),
                html.Div(style={'marginTop':'24px','display':'flex','justifyContent':'flex-end'}, children=[
                    html.Button("Fechar", id='btn-modal-config-close2', className='btn-action'),
                ]),
            ])
        ])
    ]),

    # MODAL SALVAR DRE
    html.Div(id='modal-save-overlay', style={'display':'none'}, children=[
        html.Div(className='modal-overlay', children=[
            html.Div(className='modal-box', style={'maxWidth':'400px'}, children=[
                html.Button("✕", id='btn-save-close', className='modal-close'),
                html.Div("💾 Salvar DRE", className='modal-title'),
                html.Div("Selecione o mês e ano de referência deste DRE", className='modal-sub'),
                html.Div(style={'display':'flex','gap':'12px','marginTop':'20px'}, children=[
                    html.Div(style={'flex':'1'}, children=[
                        html.Div("Mês", style={'fontFamily':"'Space Mono',monospace",'fontSize':'10px','color':'#8fa894','marginBottom':'6px','textTransform':'uppercase','letterSpacing':'2px'}),
                        dcc.Dropdown(id='save-mes',
                            options=[{'label':m,'value':str(i+1).zfill(2)} for i,m in enumerate(
                                ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
                                 'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro'])],
                            value=None, placeholder='Mês...', className='dre-dropdown'),
                    ]),
                    html.Div(style={'flex':'1'}, children=[
                        html.Div("Ano", style={'fontFamily':"'Space Mono',monospace",'fontSize':'10px','color':'#8fa894','marginBottom':'6px','textTransform':'uppercase','letterSpacing':'2px'}),
                        dcc.Dropdown(id='save-ano',
                            options=[{'label':str(y),'value':str(y)} for y in range(2023,2030)],
                            value='2025', placeholder='Ano...', className='dre-dropdown'),
                    ]),
                ]),
                html.Div(id='save-status', style={'marginTop':'14px','minHeight':'20px','fontFamily':"'Space Mono',monospace",'fontSize':'11px','color':'#3ddc84'}),
                html.Div(style={'marginTop':'20px','display':'flex','justifyContent':'flex-end','gap':'10px'}, children=[
                    html.Button("Cancelar", id='btn-save-cancel', className='btn-action',
                                style={'background':'transparent','border':'1px solid rgba(255,255,255,0.1)'}),
                    html.Button("💾 Salvar", id='btn-save-confirm', className='btn-action',
                                style={'background':'#3ddc84','color':'#0d0f0e','fontWeight':'700'}),
                ]),
            ])
        ])
    ]),

    # MODAL HISTÓRICO DRE
    html.Div(id='modal-history-overlay', style={'display':'none'}, children=[
        html.Div(className='modal-overlay', children=[
            html.Div(className='modal-box', style={'maxWidth':'520px','maxHeight':'70vh','overflowY':'auto'}, children=[
                html.Button("✕", id='btn-history-close', className='modal-close'),
                html.Div("🔍 Buscar DRE por Período", className='modal-title'),
                html.Div("Filtre e carregue um DRE salvo anteriormente", className='modal-sub'),
                html.Div(style={'display':'flex','gap':'12px','marginTop':'20px','alignItems':'flex-end'}, children=[
                    html.Div(style={'flex':'1'}, children=[
                        html.Div("Mês", style={'fontFamily':"'Space Mono',monospace",'fontSize':'10px','color':'#8fa894','marginBottom':'6px','textTransform':'uppercase','letterSpacing':'2px'}),
                        dcc.Dropdown(id='search-mes',
                            options=[{'label':m,'value':str(i+1).zfill(2)} for i,m in enumerate(
                                ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
                                 'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro'])],
                            value=None, placeholder='Todos...', className='dre-dropdown'),
                    ]),
                    html.Div(style={'flex':'1'}, children=[
                        html.Div("Ano", style={'fontFamily':"'Space Mono',monospace",'fontSize':'10px','color':'#8fa894','marginBottom':'6px','textTransform':'uppercase','letterSpacing':'2px'}),
                        dcc.Dropdown(id='search-ano',
                            options=[{'label':str(y),'value':str(y)} for y in range(2023,2030)],
                            value=None, placeholder='Todos...', className='dre-dropdown'),
                    ]),
                ]),
                html.Div(id='history-list', style={'marginTop':'20px'}),
            ])
        ])
    ]),

    # LOGIN
    html.Div(id='screen-login', className='screen-login', children=[
        html.Div(className='login-top-bar', children=[
            html.Div("GUAPO", className='login-top-brand'),
            html.Div("Sistema de Gestão Estratégica", className='login-top-tag'),
            html.Div(className='login-top-dot'),
        ]),
        html.Div(className='login-center', children=[
            html.Div(className='login-panel', children=[
                html.Div(["ACESSO", html.Span(" SEGURO")], className='login-eyebrow'),
                html.Div(["Dashboard ", html.Em("Financeiro")], className='login-headline'),
                html.P("Gestão estratégica de resultados e indicadores de performance.", className='login-desc'),
                html.Div(className='login-card', children=[
                    html.Div(className='field-group', children=[
                        html.Div("Identificação", className='field-label'),
                        dcc.Input(id="user", type="text", placeholder="Usuário do sistema", className="input-field", debounce=False),
                    ]),
                    html.Div(className='field-group', children=[
                        html.Div("Autenticação", className='field-label'),
                        dcc.Input(id="pass", type="password", placeholder="••••••••••••", className="input-field", debounce=False),
                    ]),
                    html.Button("ENTRAR NO SISTEMA →", id="btn-login", className="btn-login"),
                    html.Div(id="out-login", className="login-error"),
                ]),
                html.Div("Credenciais: admin / 123", className='login-footer-note'),
            ])
        ]),
    ]),

    # DASHBOARD
    html.Div(id='screen-dash', className='screen-dash', style={'display':'none'}, children=[
        html.Div(className='dash-layout', children=[
            # SIDEBAR
            html.Div(className='sidebar', children=[
                html.Div(className='sidebar-brand', children=[
                    html.Div(["GUAP", html.Span("O")], className='sidebar-logo'),
                    html.Div("ERP · Dashboard Financeiro", className='sidebar-sub'),
                ]),
                html.Div(className='sidebar-section', children=[
                    html.Div("Principal", className='sidebar-section-label'),
                    html.Div(id='nav-ind',    children=[html.Span("🏠", className='nav-icon'), "Início"],              className='nav-item active'),
                    html.Div(id='nav-charts', children=[html.Span("📈", className='nav-icon'), "Gráficos"],         className='nav-item'),
                    html.Div(id='nav-dre',    children=[html.Span("📋", className='nav-icon'), "DRE Completo"],     className='nav-item'),
                    html.Div(id='nav-txn',    children=[html.Span("💳", className='nav-icon'), "Transações"],       className='nav-item'),
                    html.Div(id='nav-conc',   children=[html.Span("🏦", className='nav-icon'), "Conciliação Bancária"], className='nav-item'),
                ]),
                html.Div(className='sidebar-section', children=[
                    html.Div("Ferramentas", className='sidebar-section-label'),
                    html.Div(id='nav-export',   children=[html.Span("⬇️",  className='nav-icon'), "Exportar Excel"],       className='nav-item'),
                    html.Div(id='nav-template', children=[html.Span("📄",  className='nav-icon'), "Baixar Template CSV"],  className='nav-item'),
                    html.Div(id='nav-save-dre', children=[html.Span("💾",  className='nav-icon'), "Salvar DRE"],            className='nav-item'),
                    html.Div(id='nav-history',  children=[html.Span("🔍",  className='nav-icon'), "Buscar DRE"],            className='nav-item'),
                    html.Div(id='nav-config',   children=[html.Span("⚙️",  className='nav-icon'), "Configurações"],        className='nav-item'),
                    html.Div(id='nav-logout',   children=[html.Span("🚪",  className='nav-icon'), "Sair"],                 className='nav-item'),
                ]),
                html.Div(className='sidebar-bottom', children=[
                    html.Div(className='sidebar-user', children=[
                        html.Div("A", className='user-avatar'),
                        html.Div([html.Div("Administrador", className='user-name'), html.Div("Acesso Total", className='user-role')]),
                    ])
                ]),
            ]),

            # MAIN
            html.Div(className='dash-main', children=[
                html.Div(className='topbar', children=[
                    html.Div([
                        html.Div(id='topbar-title', children="Rede Guapo", className='topbar-title'),
                        html.Div(id='topbar-subtitle', children="ERP Financeiro · Gestão Estratégica de Resultados", className='topbar-subtitle'),
                    ]),
                    html.Div(className='topbar-actions', children=[
                        html.Div(id='topbar-date', className='topbar-date'),
                        html.Div([html.Div(className='status-dot'), "SISTEMA ATIVO"], className='status-chip'),
                    ]),
                ]),
                html.Div(id='main-content', className='main-content'),
                html.Div(id='dre-tab-wrapper',  className='main-content', style={'display':'none'}, children=[_build_dre_tab()]),
                html.Div(id='conc-tab-wrapper', className='main-content', style={'display':'none'}, children=[
                    html.Div(className='section-header', children=[
                        html.Div([
                            html.Div("Extrato Bancário", className='section-eyebrow'),
                            html.Div("Conciliação Bancária", className='section-title'),
                        ]),
                        html.Div("Importe o extrato Excel para classificar automaticamente as transações",
                                 style={'fontSize':'12px','color':'#4d5e52','fontFamily':"'Space Mono',monospace"}),
                    ]),
                    _build_conciliacao_tab(),
                ]),
            ]),
        ]),
    ]),
])

# ═══════════════════════════════════════════════════════════════════════════════
# CALLBACKS
# ═══════════════════════════════════════════════════════════════════════════════

@app.callback(
    Output('screen-login','style'),
    Output('screen-dash','style'),
    Output('out-login','children'),
    Input('btn-login','n_clicks'),
    State('user','value'), State('pass','value'),
    prevent_initial_call=True,
)
def do_login(n, u, p):
    if u == 'admin' and p == '123':
        return {'display':'none'}, {'display':'block'}, ''
    return {'display':'flex'}, {'display':'none'}, '✕  Credenciais inválidas'

@app.callback(
    Output('screen-login','style', allow_duplicate=True),
    Output('screen-dash','style',  allow_duplicate=True),
    Input('nav-logout','n_clicks'),
    prevent_initial_call=True,
)
def do_logout(_):
    return {'display':'flex'}, {'display':'none'}

@app.callback(Output('topbar-date','children'), Input('clock-tick','n_intervals'))
def update_clock(_):
    return datetime.datetime.now().strftime('%d/%m/%Y  %H:%M')

@app.callback(
    Output('active-tab','data'),
    Output('nav-ind','className'),
    Output('nav-charts','className'),
    Output('nav-dre','className'),
    Output('nav-txn','className'),
    Output('nav-conc','className'),
    Output('topbar-title','children'),
    Output('topbar-subtitle','children'),
    Output('main-content','style'),
    Output('dre-tab-wrapper','style'),
    Output('conc-tab-wrapper','style'),
    Input('nav-ind','n_clicks'),
    Input('nav-charts','n_clicks'),
    Input('nav-dre','n_clicks'),
    Input('nav-txn','n_clicks'),
    Input('nav-conc','n_clicks'),
    Input('csv-store','data'),
    prevent_initial_call=True,
)
def nav_click(n1, n2, n3, n4, n5, csv_data):
    triggered = ctx.triggered_id
    show = {'display':'block'}
    hide = {'display':'none'}

    if triggered == 'csv-store':
        if not csv_data:
            raise dash.exceptions.PreventUpdate
        return ('tab-dre',
                'nav-item','nav-item','nav-item active','nav-item','nav-item',
                'DRE Completo','CSV importado · campos preenchidos automaticamente',
                hide, show, hide)
    tabs = {
        'nav-ind':    ('tab-home',   'Rede Guapo',              'ERP Financeiro · Gestão Estratégica de Resultados'),
        'nav-charts': ('tab-charts', 'Análise Gráfica',         'Visualização de dados financeiros'),
        'nav-dre':    ('tab-dre',    'DRE Completo',            'Demonstração do Resultado do Exercício · Panelas do Guapo'),
        'nav-txn':    ('tab-txn',    'Transações Recentes',     'Histórico de entradas e saídas'),
        'nav-conc':   ('tab-conc',   'Conciliação Bancária',    'Extrato bancário · Classificação automática'),
    }
    tab, title, sub = tabs.get(triggered, ('tab-home','Panelas do Guapo',''))
    cls = {k: 'nav-item active' if k == triggered else 'nav-item' for k in tabs}
    is_dre  = tab == 'tab-dre'
    is_conc = tab == 'tab-conc'
    return (tab,
            cls['nav-ind'], cls['nav-charts'], cls['nav-dre'], cls['nav-txn'], cls['nav-conc'],
            title, sub,
            hide if (is_dre or is_conc) else show,
            show if is_dre else hide,
            show if is_conc else hide)

# MODAL CONFIG
@app.callback(
    Output('modal-config-overlay','style'),
    Input('nav-config','n_clicks'),
    Input('btn-modal-config-close','n_clicks'),
    Input('btn-modal-config-close2','n_clicks'),
    prevent_initial_call=True,
)
def toggle_modal_config(*_):
    return {'display':'block'} if ctx.triggered_id == 'nav-config' else {'display':'none'}

@app.callback(Output('toggle-notif','className'), Input('toggle-notif','n_clicks'), State('toggle-notif','className'), prevent_initial_call=True)
def _t1(n, cls): return 'toggle off' if cls and 'on' in cls else 'toggle on'
@app.callback(Output('toggle-dark','className'), Input('toggle-dark','n_clicks'), State('toggle-dark','className'), prevent_initial_call=True)
def _t2(n, cls): return 'toggle off' if cls and 'on' in cls else 'toggle on'
@app.callback(Output('toggle-autoexp','className'), Input('toggle-autoexp','n_clicks'), State('toggle-autoexp','className'), prevent_initial_call=True)
def _t3(n, cls): return 'toggle off' if cls and 'on' in cls else 'toggle on'
@app.callback(Output('toggle-auto','className'), Input('toggle-auto','n_clicks'), State('toggle-auto','className'), prevent_initial_call=True)
def _t4(n, cls): return 'toggle off' if cls and 'on' in cls else 'toggle on'

@app.callback(
    Output('download-excel','data'),
    Output('toast','children', allow_duplicate=True),
    Output('toast','style', allow_duplicate=True),
    Input('nav-export','n_clicks'),
    State('dre-store','data'),
    prevent_initial_call=True,
)
def export_excel(n, dre_data):
    if not n: raise dash.exceptions.PreventUpdate
    data = gerar_excel(dre_data or {})
    fname = f"GUAPO_DRE_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    toast = html.Div([html.Span("✓", style={'color':'#3ddc84','fontWeight':'700','fontSize':'16px'}),
                      html.Span(f" Excel exportado — {fname}", style={'fontFamily':"'Space Mono',monospace",'fontSize':'11px'})], className='toast')
    return dcc.send_bytes(data, fname), toast, {'display':'block'}

@app.callback(
    Output('download-template','data'),
    Output('toast','children', allow_duplicate=True),
    Output('toast','style', allow_duplicate=True),
    Input('nav-template','n_clicks'),
    prevent_initial_call=True,
)
def download_template(n):
    if not n: raise dash.exceptions.PreventUpdate
    toast = html.Div([html.Span("📄", style={'fontSize':'16px'}),
                      html.Span(" Template CSV baixado", style={'fontFamily':"'Space Mono',monospace",'fontSize':'11px'})], className='toast')
    return dcc.send_bytes(gerar_template_csv(), "GUAPO_DRE_template.csv"), toast, {'display':'block'}

@app.callback(
    Output('csv-store', 'data'),
    Output('toast', 'children', allow_duplicate=True),
    Output('toast', 'style', allow_duplicate=True),
    Input('upload-data', 'contents'),
    State('upload-data', 'filename'),
    prevent_initial_call=True,
)
def parse_upload(contents, filename):
    if not contents: raise dash.exceptions.PreventUpdate
    _, content_string = contents.split(',')
    try:
        decoded = base64.b64decode(content_string).decode('utf-8')
    except Exception:
        decoded = base64.b64decode(content_string).decode('latin-1')
    result, n_campos = parse_csv_to_dre(decoded)
    if not result:
        toast = html.Div([html.Span("⚠", style={'color':'#ff5c5c','fontWeight':'700','fontSize':'16px'}),
                          html.Span(" Nenhum campo reconhecido — baixe o template CSV", style={'fontFamily':"'Space Mono',monospace",'fontSize':'11px'})], className='toast')
        return dash.no_update, toast, {'display':'block'}
    toast = html.Div([html.Span("✓", style={'color':'#3ddc84','fontWeight':'700','fontSize':'16px'}),
                      html.Span(f" {filename} importado · {n_campos} campos preenchidos", style={'fontFamily':"'Space Mono',monospace",'fontSize':'11px'})], className='toast')
    return result, toast, {'display':'block'}

_ALL_DRE_FIELD_IDS = [
    'dre-venda-vista','dre-venda-prazo',
    'dre-g-bolos','dre-g-buffet','dre-g-cafes','dre-g-lanches','dre-g-porcoes',
    'dre-g-salgados','dre-g-drinks','dre-g-conv','dre-g-cigarros','dre-g-bebidas',
    'dre-g-picole','dre-g-estcoz','dre-g-bichos','dre-g-carnes',
    'dre-or-juros','dre-or-alug','dre-or-outras','dre-or-fundos','dre-or-bonif',
    'dre-c-bebidas','dre-c-bolos','dre-c-carnes','dre-c-cigarros','dre-c-conv',
    'dre-c-estcoz','dre-c-picoles','dre-c-salgados','dre-c-insumos','dre-c-embal',
    'dre-c-recarga','dre-c-bichos','dre-c-buffet',
    'dre-cf-energia','dre-cf-tel','dre-cf-agua','dre-cf-iptu','dre-cf-sal',
    'dre-cf-enc','dre-cf-imp','dre-cf-seg','dre-cf-diar',
    'dre-cv-sist','dre-cv-terc','dre-cv-exped','dre-cv-honor','dre-cv-manut',
    'dre-cv-viag','dre-cv-taxas','dre-cv-unif','dre-cv-limp','dre-cv-alug',
    'dre-cv-caixa','dre-cv-mora','dre-cv-banco','dre-cv-cart','dre-cv-brindes',
    'dre-cv-utens','dre-cv-veic','dre-cv-segpred','dre-cv-gerais','dre-cv-gas',
    'dre-cv-estoque','dre-cv-comiss',
]

def _field_id_to_key(fid):
    return fid.replace('dre-', '').replace('-', '_')

@app.callback(
    [Output(fid, 'value') for fid in _ALL_DRE_FIELD_IDS],
    Input('csv-store', 'data'),
    prevent_initial_call=True,
)
def fill_dre_from_csv(csv_data):
    if not csv_data: raise dash.exceptions.PreventUpdate
    def fmt_brl(v):
        if v is None: return None
        try:
            return f'{float(v):,.2f}'.replace(',','X').replace('.', ',').replace('X','.')
        except: return None
    return [fmt_brl(csv_data.get(_field_id_to_key(fid))) for fid in _ALL_DRE_FIELD_IDS]

@app.callback(
    Output('dre-store', 'data'),
    Input('dre-venda-vista','value'), Input('dre-venda-prazo','value'),
    Input('dre-or-juros','value'), Input('dre-or-alug','value'),
    Input('dre-or-outras','value'), Input('dre-or-fundos','value'), Input('dre-or-bonif','value'),
    Input('dre-g-bolos','value'), Input('dre-g-buffet','value'), Input('dre-g-cafes','value'),
    Input('dre-g-lanches','value'), Input('dre-g-porcoes','value'), Input('dre-g-salgados','value'),
    Input('dre-g-drinks','value'), Input('dre-g-conv','value'), Input('dre-g-cigarros','value'),
    Input('dre-g-bebidas','value'), Input('dre-g-picole','value'), Input('dre-g-estcoz','value'),
    Input('dre-g-bichos','value'), Input('dre-g-carnes','value'),
    Input('dre-c-bebidas','value'), Input('dre-c-bolos','value'), Input('dre-c-carnes','value'),
    Input('dre-c-cigarros','value'), Input('dre-c-conv','value'), Input('dre-c-estcoz','value'),
    Input('dre-c-picoles','value'), Input('dre-c-salgados','value'), Input('dre-c-insumos','value'),
    Input('dre-c-embal','value'), Input('dre-c-recarga','value'), Input('dre-c-bichos','value'),
    Input('dre-c-buffet','value'),
    Input('dre-cf-energia','value'), Input('dre-cf-tel','value'), Input('dre-cf-agua','value'),
    Input('dre-cf-iptu','value'), Input('dre-cf-sal','value'), Input('dre-cf-enc','value'),
    Input('dre-cf-imp','value'), Input('dre-cf-seg','value'), Input('dre-cf-diar','value'),
    Input('dre-cv-sist','value'), Input('dre-cv-terc','value'), Input('dre-cv-exped','value'),
    Input('dre-cv-honor','value'), Input('dre-cv-manut','value'), Input('dre-cv-viag','value'),
    Input('dre-cv-taxas','value'), Input('dre-cv-unif','value'), Input('dre-cv-limp','value'),
    Input('dre-cv-alug','value'), Input('dre-cv-caixa','value'), Input('dre-cv-mora','value'),
    Input('dre-cv-banco','value'), Input('dre-cv-cart','value'), Input('dre-cv-brindes','value'),
    Input('dre-cv-utens','value'), Input('dre-cv-veic','value'), Input('dre-cv-segpred','value'),
    Input('dre-cv-gerais','value'), Input('dre-cv-gas','value'), Input('dre-cv-estoque','value'),
    Input('dre-cv-comiss','value'),
    prevent_initial_call=False,
)
def save_dre_store(*vals):
    keys = [
        'venda_vista','venda_prazo','or_juros','or_alug','or_outras','or_fundos','or_bonif',
        'g_bolos','g_buffet','g_cafes','g_lanches','g_porcoes','g_salgados','g_drinks','g_conv',
        'g_cigarros','g_bebidas','g_picole','g_estcoz','g_bichos','g_carnes',
        'c_bebidas','c_bolos','c_carnes','c_cigarros','c_conv','c_estcoz','c_picoles','c_salgados',
        'c_insumos','c_embal','c_recarga','c_bichos','c_buffet',
        'cf_energia','cf_tel','cf_agua','cf_iptu','cf_sal','cf_enc','cf_imp','cf_seg','cf_diar',
        'cv_sist','cv_terc','cv_exped','cv_honor','cv_manut','cv_viag','cv_taxas','cv_unif',
        'cv_limp','cv_alug','cv_caixa','cv_mora','cv_banco','cv_cart','cv_brindes','cv_utens',
        'cv_veic','cv_segpred','cv_gerais','cv_gas','cv_estoque','cv_comiss',
    ]
    def to_float(v):
        if v is None or v == '': return 0.0
        if isinstance(v, (int, float)): return float(v)
        s = str(v).strip().replace('R$','').replace(' ','').replace('.','').replace(',','.')
        try: return float(s)
        except: return 0.0
    return {k: to_float(v) for k, v in zip(keys, vals)}

def _parse(v):
    if v is None or v == '': return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace('R$','').replace(' ','').replace('.','').replace(',','.')
    try: return float(s)
    except: return 0.0

@app.callback(
    Output('res-receita-total', 'children'),
    Input('dre-venda-vista','value'), Input('dre-venda-prazo','value'),
    Input('dre-or-juros','value'), Input('dre-or-alug','value'),
    Input('dre-or-outras','value'), Input('dre-or-fundos','value'), Input('dre-or-bonif','value'),
    prevent_initial_call=False,
)
def calc_receita_total(*vals):
    total = sum(_parse(v) for v in vals if v is not None)
    if total == 0: return 'R$ 0,00'
    return f"R$ {total:,.2f}".replace(',','X').replace('.', ',').replace('X','.')

@app.callback(
    Output('res-margem', 'children'),
    Input('dre-venda-vista','value'), Input('dre-venda-prazo','value'),
    Input('dre-or-juros','value'), Input('dre-or-alug','value'),
    Input('dre-or-outras','value'), Input('dre-or-fundos','value'), Input('dre-or-bonif','value'),
    Input('dre-c-bebidas','value'), Input('dre-c-bolos','value'), Input('dre-c-carnes','value'),
    Input('dre-c-cigarros','value'), Input('dre-c-conv','value'), Input('dre-c-estcoz','value'),
    Input('dre-c-picoles','value'), Input('dre-c-salgados','value'), Input('dre-c-insumos','value'),
    Input('dre-c-embal','value'), Input('dre-c-recarga','value'), Input('dre-c-bichos','value'),
    Input('dre-c-buffet','value'),
    prevent_initial_call=False,
)
def calc_margem(*vals):
    receita = sum(_parse(v) for v in vals[:7] if v is not None)
    compras = sum(_parse(v) for v in vals[7:] if v is not None)
    margem  = receita - compras
    if receita == 0 and compras == 0: return '—'
    return f"R$ {margem:,.2f}".replace(',','X').replace('.', ',').replace('X','.')

_REC_IDS  = ['dre-venda-vista','dre-venda-prazo','dre-or-juros','dre-or-alug','dre-or-outras','dre-or-fundos','dre-or-bonif']
_COMP_IDS = ['dre-c-bebidas','dre-c-bolos','dre-c-carnes','dre-c-cigarros','dre-c-conv','dre-c-estcoz','dre-c-picoles','dre-c-salgados','dre-c-insumos','dre-c-embal','dre-c-recarga','dre-c-bichos','dre-c-buffet']
_CF_IDS   = ['dre-cf-energia','dre-cf-tel','dre-cf-agua','dre-cf-iptu','dre-cf-sal','dre-cf-enc','dre-cf-imp','dre-cf-seg','dre-cf-diar']
_CV_IDS   = ['dre-cv-sist','dre-cv-terc','dre-cv-exped','dre-cv-honor','dre-cv-manut','dre-cv-viag','dre-cv-taxas','dre-cv-unif','dre-cv-limp','dre-cv-alug','dre-cv-caixa','dre-cv-mora','dre-cv-banco','dre-cv-cart','dre-cv-brindes','dre-cv-utens','dre-cv-veic','dre-cv-segpred','dre-cv-gerais','dre-cv-gas','dre-cv-estoque','dre-cv-comiss']

@app.callback(
    Output('res-lucro-bruto','children'),
    Output('res-operacional', 'children'),
    Output('res-liquido',     'children'),
    [Input(i,'value') for i in _REC_IDS + _COMP_IDS + _CF_IDS + _CV_IDS],
    prevent_initial_call=False,
)
def calc_resultados(*vals):
    nr = len(_REC_IDS); nc = len(_COMP_IDS); nf = len(_CF_IDS)
    receita     = sum(_parse(v) for v in vals[:nr] if v is not None)
    compras     = sum(_parse(v) for v in vals[nr:nr+nc] if v is not None)
    cf          = sum(_parse(v) for v in vals[nr+nc:nr+nc+nf] if v is not None)
    cv          = sum(_parse(v) for v in vals[nr+nc+nf:] if v is not None)
    lucro_bruto = (receita - compras) - cf
    operacional = lucro_bruto - cv
    def fmt(v):
        if receita == 0 and compras == 0 and cf == 0 and cv == 0: return '—'
        return f"R$ {v:,.2f}".replace(',','X').replace('.', ',').replace('X','.')
    return fmt(lucro_bruto), fmt(operacional), fmt(operacional)

@app.callback(Output('fig-debitos', 'figure'), Input('dre-store', 'data'))
def fig_debitos(d):
    if not d: d = {}
    cf_items = {'Energia Elétrica': d.get('cf_energia',0), 'Telefone': d.get('cf_tel',0),
                'Água': d.get('cf_agua',0), 'IPTU': d.get('cf_iptu',0),
                'Salários/13º/Férias': d.get('cf_sal',0), 'Encargos Sociais': d.get('cf_enc',0),
                'Impostos': d.get('cf_imp',0), 'Seguro Vida': d.get('cf_seg',0), 'Diárias': d.get('cf_diar',0)}
    cv_items = {'Sistemas': d.get('cv_sist',0), 'Serv. Terceiros': d.get('cv_terc',0),
                'Expediente': d.get('cv_exped',0), 'Honorários': d.get('cv_honor',0),
                'Manutenção': d.get('cv_manut',0), 'Viagens': d.get('cv_viag',0),
                'Taxas': d.get('cv_taxas',0), 'Uniformes': d.get('cv_unif',0),
                'Limpeza': d.get('cv_limp',0), 'Aluguéis': d.get('cv_alug',0),
                'Caixa': d.get('cv_caixa',0), 'Juros Mora': d.get('cv_mora',0),
                'Tarifas Banco': d.get('cv_banco',0), 'Tarifas Cartão': d.get('cv_cart',0),
                'Brindes': d.get('cv_brindes',0), 'Utensílios': d.get('cv_utens',0),
                'Veículos': d.get('cv_veic',0), 'Seg. Edificações': d.get('cv_segpred',0),
                'Desp. Gerais': d.get('cv_gerais',0), 'Gás GLP': d.get('cv_gas',0),
                'Estoque': d.get('cv_estoque',0), 'Comissões': d.get('cv_comiss',0)}
    compras_items = {'Bebidas': d.get('c_bebidas',0), 'Bolos/Tortas': d.get('c_bolos',0),
                     'Carnes': d.get('c_carnes',0), 'Cigarros': d.get('c_cigarros',0),
                     'Conveniência': d.get('c_conv',0), 'Est. Cozinha': d.get('c_estcoz',0),
                     'Picolés': d.get('c_picoles',0), 'Salgados': d.get('c_salgados',0),
                     'Insumos': d.get('c_insumos',0), 'Embalagens': d.get('c_embal',0),
                     'Recarga': d.get('c_recarga',0), 'Bichos/Brinq.': d.get('c_bichos',0), 'Buffet': d.get('c_buffet',0)}
    total_cf = sum(cf_items.values()); total_cv = sum(cv_items.values())
    total_deb = total_cf + total_cv; total_comp = sum(compras_items.values())
    total_geral = total_deb + total_comp
    fig = go.Figure()
    for nome, val in cf_items.items():
        if val > 0:
            fig.add_trace(go.Bar(name=nome, x=['Custos Fixos'], y=[val], marker_color='rgba(255,92,92,0.75)', marker_line_width=0, showlegend=True, hovertemplate=f'<b>{nome}</b><br>R$ %{{y:,.2f}}<extra></extra>'))
    for nome, val in cv_items.items():
        if val > 0:
            fig.add_trace(go.Bar(name=nome, x=['Custos Variáveis'], y=[val], marker_color='rgba(255,150,50,0.75)', marker_line_width=0, showlegend=True, hovertemplate=f'<b>{nome}</b><br>R$ %{{y:,.2f}}<extra></extra>'))
    fig.add_trace(go.Bar(name='Total Débitos', x=['Total Débitos'], y=[total_deb], marker_color='rgba(255,92,92,0.9)', marker_line_width=0, showlegend=False, hovertemplate='<b>Total Débitos</b><br>R$ %{y:,.2f}<extra></extra>'))
    for nome, val in compras_items.items():
        if val > 0:
            fig.add_trace(go.Bar(name=nome, x=['Compras Mercadorias'], y=[val], marker_color='rgba(201,168,76,0.75)', marker_line_width=0, showlegend=True, hovertemplate=f'<b>{nome}</b><br>R$ %{{y:,.2f}}<extra></extra>'))
    fig.add_trace(go.Bar(name='Total Geral', x=['Total Geral'], y=[total_geral], marker_color='rgba(92,158,255,0.9)', marker_line_width=0, showlegend=False, hovertemplate='<b>Total Geral</b><br>R$ %{y:,.2f}<extra></extra>'))
    layout = dict(CHART_LAYOUT); layout['barmode']='stack'; layout['margin']=dict(l=16,r=16,t=16,b=80); layout['showlegend']=False
    layout['xaxis']=dict(gridcolor='rgba(255,255,255,0.04)', tickfont=dict(size=11,color='#e8ede9',family='Sora'))
    fig.update_layout(**layout)
    for label, val in [('Custos Fixos',total_cf),('Custos Variáveis',total_cv),('Total Débitos',total_deb),('Compras Mercadorias',total_comp),('Total Geral',total_geral)]:
        if val > 0:
            fig.add_annotation(x=label, y=val, text=f"R$ {val:,.0f}".replace(',','X').replace('.', ',').replace('X','.'), showarrow=False, yshift=10, font=dict(size=11,color='#e8ede9',family='Space Mono'))
    return fig

@app.callback(Output('fig-receitas', 'figure'), Input('dre-store', 'data'))
def fig_receitas(d):
    if not d: d = {}
    vendas_merc  = {'Vendas à Vista': d.get('venda_vista',0), 'Vendas a Prazo': d.get('venda_prazo',0)}
    vendas_grupo = {'Bolos e Tortas': d.get('g_bolos',0), 'Buffet': d.get('g_buffet',0),
                    'Cafés e Sucos': d.get('g_cafes',0), 'Lanches': d.get('g_lanches',0),
                    'Porções': d.get('g_porcoes',0), 'Salgados': d.get('g_salgados',0),
                    'Drinks': d.get('g_drinks',0), 'Conveniência': d.get('g_conv',0),
                    'Cigarros': d.get('g_cigarros',0), 'Bebidas': d.get('g_bebidas',0),
                    'Picolé': d.get('g_picole',0), 'Est. Cozinha': d.get('g_estcoz',0),
                    'Bichos/Brinq.': d.get('g_bichos',0), 'Carnes': d.get('g_carnes',0)}
    outras_rec = {'Juros Recebidos': d.get('or_juros',0), 'Aluguéis': d.get('or_alug',0),
                  'Outras Receitas': d.get('or_outras',0), 'Fundos Invest.': d.get('or_fundos',0),
                  'Bonificações': d.get('or_bonif',0)}
    total_merc = sum(vendas_merc.values()); total_grupo = sum(vendas_grupo.values())
    total_outras = sum(outras_rec.values()); total_rec = total_merc + total_grupo + total_outras
    fig = go.Figure()
    for nome, val in vendas_merc.items():
        if val > 0: fig.add_trace(go.Bar(name=nome, x=['Vendas Mercadorias'], y=[val], marker_color='rgba(61,220,132,0.8)', marker_line_width=0, showlegend=True, hovertemplate=f'<b>{nome}</b><br>R$ %{{y:,.2f}}<extra></extra>'))
    for nome, val in vendas_grupo.items():
        if val > 0: fig.add_trace(go.Bar(name=nome, x=['Vendas por Grupo'], y=[val], marker_color='rgba(61,180,100,0.75)', marker_line_width=0, showlegend=True, hovertemplate=f'<b>{nome}</b><br>R$ %{{y:,.2f}}<extra></extra>'))
    for nome, val in outras_rec.items():
        if val > 0: fig.add_trace(go.Bar(name=nome, x=['Outras Receitas'], y=[val], marker_color='rgba(92,158,255,0.8)', marker_line_width=0, showlegend=True, hovertemplate=f'<b>{nome}</b><br>R$ %{{y:,.2f}}<extra></extra>'))
    fig.add_trace(go.Bar(name='Receita Total', x=['Receita Total'], y=[total_rec], marker_color='rgba(61,220,132,0.95)', marker_line_width=0, showlegend=False, hovertemplate='<b>Receita Total</b><br>R$ %{y:,.2f}<extra></extra>'))
    layout = dict(CHART_LAYOUT); layout['barmode']='stack'; layout['margin']=dict(l=16,r=16,t=16,b=80); layout['showlegend']=False
    layout['xaxis']=dict(gridcolor='rgba(255,255,255,0.04)', tickfont=dict(size=11,color='#e8ede9',family='Sora'))
    fig.update_layout(**layout)
    for label, val in [('Vendas Mercadorias',total_merc),('Vendas por Grupo',total_grupo),('Outras Receitas',total_outras),('Receita Total',total_rec)]:
        if val > 0:
            fig.add_annotation(x=label, y=val, text=f"R$ {val:,.0f}".replace(',','X').replace('.', ',').replace('X','.'), showarrow=False, yshift=10, font=dict(size=11,color='#e8ede9',family='Space Mono'))
    return fig

@app.callback(
    Output('upload-status-text','children'),
    Output('upload-status-text','className'),
    Output('csv-banner','children'),
    Output('csv-banner','style'),
    Input('csv-store','data'),
    prevent_initial_call=False,
)
def update_csv_banner(csv_data):
    if not csv_data:
        return "AGUARDANDO CSV", "upload-status", [
            html.Div([html.Strong("Como importar via CSV: "), "Baixe o template, preencha e importe."]),
            html.Div([html.Strong("Formato: "), "duas colunas — campo e valor."]),
        ], {}
    n = len(csv_data)
    return f"✓ CSV CARREGADO", "upload-status loaded", [
        html.Div([html.Strong(f"✓ CSV importado · {n} campos preenchidos — "), "você pode editar manualmente."]),
    ], {'borderColor':'rgba(61,220,132,0.3)','background':'rgba(61,220,132,0.06)'}

def make_fig_main():
    fig = go.Figure()
    fig.add_trace(go.Bar(name='Receita', x=MESES, y=RECEITA, marker_color='rgba(61,220,132,0.7)', marker_line_width=0))
    fig.add_trace(go.Bar(name='Despesa', x=MESES, y=DESPESA, marker_color='rgba(255,92,92,0.6)', marker_line_width=0))
    fig.add_trace(go.Scatter(name='Lucro', x=MESES, y=LUCRO, mode='lines+markers',
                             line=dict(color='#e8c97a',width=2.5), marker=dict(size=6,color='#e8c97a')))
    fig.update_layout(**CHART_LAYOUT, barmode='group')
    return fig

def _home_feature_card(icon, title, desc, nav_id, badge=None, badge_cls="up"):
    return html.Div(style={
        'background':'#131714','border':'1px solid rgba(255,255,255,0.07)',
        'borderRadius':'12px','padding':'28px 26px','display':'flex',
        'flexDirection':'column','gap':'12px','transition':'all 0.3s',
        'cursor':'default','position':'relative','overflow':'hidden',
    }, children=[
        html.Div(style={
            'position':'absolute','top':0,'left':0,'right':0,'height':'2px',
            'background':'linear-gradient(90deg,var(--green),transparent)','opacity':'0.4',
        }),
        html.Div(style={'display':'flex','alignItems':'center','justifyContent':'space-between'}, children=[
            html.Div(icon, style={'fontSize':'26px'}),
            html.Div(badge, className=f'kpi-badge {badge_cls}') if badge else html.Div(),
        ]),
        html.Div(title, style={
            'fontFamily':"'Playfair Display',serif",'fontSize':'17px',
            'fontWeight':'700','color':'#e8ede9','lineHeight':'1.2',
        }),
        html.Div(desc, style={
            'fontSize':'12px','color':'#8fa894','lineHeight':'1.6',
            'fontFamily':"'Sora',sans-serif",'flex':'1',
        }),
        html.Button("Acessar →", id=f'home-btn-{nav_id}',
            style={
                'marginTop':'8px','padding':'10px 20px','borderRadius':'6px',
                'border':'1px solid rgba(61,220,132,0.3)',
                'background':'rgba(61,220,132,0.07)','color':'#3ddc84',
                'fontFamily':"'Space Mono',monospace",'fontSize':'10px',
                'fontWeight':'700','letterSpacing':'2px','textTransform':'uppercase',
                'cursor':'pointer','width':'fit-content',
                'transition':'all 0.25s',
            }
        ),
    ])

@app.callback(Output('main-content','children'), Input('active-tab','data'))
def render_content(tab):
    if tab in ('tab-home', None):
        return html.Div([
            # Hero
            html.Div(style={'marginBottom':'36px'}, children=[
                html.Div("BEM-VINDO AO SISTEMA", style={
                    'fontFamily':"'Space Mono',monospace",'fontSize':'10px','color':'#3ddc84',
                    'textTransform':'uppercase','letterSpacing':'4px','marginBottom':'10px',
                }),
                html.Div(["REDE ", html.Em("GUAPO", style={'color':'#3ddc84'})], style={
                    'fontFamily':"'Playfair Display',serif",'fontSize':'42px',
                    'fontWeight':'900','color':'#e8ede9','letterSpacing':'-1px','lineHeight':'1',
                    'marginBottom':'10px',
                }),
                html.Div("ERP Financeiro · Gestão Estratégica de Resultados", style={
                    'fontFamily':"'Space Mono',monospace",'fontSize':'11px',
                    'color':'#4d5e52','letterSpacing':'1px',
                }),
            ]),

            # Cards de funcionalidades — grid 3 colunas
            html.Div(style={
                'display':'grid','gridTemplateColumns':'repeat(3,1fr)',
                'gap':'18px','marginBottom':'32px',
            }, children=[
                _home_feature_card("📋", "DRE Completo",
                    "Lançamento completo da Demonstração do Resultado do Exercício. Preencha receitas, compras, custos fixos e variáveis. Cálculos automáticos de margem, lucro bruto, operacional e líquido.",
                    "dre", badge="PRINCIPAL", badge_cls="up"),
                _home_feature_card("📊", "Indicadores & KPIs",
                    "Visualize os indicadores financeiros reais calculados diretamente do DRE: Receita Total, Margem de Contribuição, Lucro Bruto, Resultado Líquido e muito mais.",
                    "dre", badge="REAL-TIME", badge_cls="up"),
                _home_feature_card("📈", "Gráficos & Análises",
                    "Visualização completa da estrutura de custos e composição de receitas em gráficos interativos empilhados por categoria, gerados automaticamente dos dados do DRE.",
                    "charts", badge="VISUAL", badge_cls="blue"),
                _home_feature_card("🏦", "Conciliação Bancária",
                    "Importe o extrato bancário (.xlsx) do Sicredi e classifique automaticamente todas as transações por tipo: PIX Recebido, Boletos, Cartão, Débito Automático e mais.",
                    "conc", badge="AUTOMÁTICO", badge_cls="wait"),
                _home_feature_card("💳", "Transações Recentes",
                    "Histórico de entradas e saídas financeiras com categorias, variações e status. Acompanhe os lançamentos mais recentes do período.",
                    "txn", badge="HISTÓRICO", badge_cls="wait"),
                _home_feature_card("📥", "Exportar & Importar",
                    "Exporte o DRE formatado em Excel profissional, baixe o template CSV para importação em lote e salve DREs por período para consulta no histórico.",
                    "export", badge="FERRAMENTAS", badge_cls="blue"),
            ]),

            # Rodapé informativo
            html.Div(style={
                'background':'rgba(61,220,132,0.04)','border':'1px solid rgba(61,220,132,0.1)',
                'borderRadius':'10px','padding':'18px 24px',
                'display':'flex','alignItems':'center','gap':'16px',
            }, children=[
                html.Div("💡", style={'fontSize':'20px'}),
                html.Div([
                    html.Div("Como começar",style={
                        'fontFamily':"'Space Mono',monospace",'fontSize':'10px',
                        'color':'#3ddc84','textTransform':'uppercase','letterSpacing':'2px','marginBottom':'4px',
                    }),
                    html.Div(
                        "Acesse o menu DRE Completo para lançar os dados do mês. Os Indicadores serão calculados automaticamente. Use Conciliação Bancária para importar o extrato do Sicredi.",
                        style={'fontSize':'12px','color':'#8fa894','lineHeight':'1.7'},
                    ),
                ]),
            ]),
        ])

    if tab == 'tab-ind':
        # Redireciona para home — não deve chegar aqui normalmente
        return html.Div()

    elif tab == 'tab-charts':
        return html.Div([
            html.Div(className='section-header', children=[
                html.Div([html.Div("Visualização de Dados", className='section-eyebrow'), html.Div("Análise Gráfica — DRE", className='section-title')]),
            ]),
            html.Div(style={'marginBottom':'24px'}, children=[
                html.Div(style={'background':'#131714','border':'1px solid rgba(255,255,255,0.07)','borderRadius':'10px','overflow':'hidden'}, children=[
                    html.Div(style={'padding':'16px 22px 12px','borderBottom':'1px solid rgba(255,255,255,0.06)','background':'#1a1f1c'}, children=[
                        html.Div("ESTRUTURA DE CUSTOS", style={'fontFamily':"'Space Mono',monospace",'fontSize':'9px','color':'#ff5c5c','textTransform':'uppercase','letterSpacing':'3px','marginBottom':'3px'}),
                        html.Div("Custos Fixos + Variáveis + Compras", style={'fontFamily':"'Playfair Display',serif",'fontSize':'15px','fontWeight':'700','color':'#e8ede9'}),
                    ]),
                    dcc.Graph(id='fig-debitos', config={'displayModeBar':False}, style={'height':'340px'}),
                ]),
            ]),
            html.Div(children=[
                html.Div(style={'background':'#131714','border':'1px solid rgba(255,255,255,0.07)','borderRadius':'10px','overflow':'hidden'}, children=[
                    html.Div(style={'padding':'16px 22px 12px','borderBottom':'1px solid rgba(255,255,255,0.06)','background':'#1a1f1c'}, children=[
                        html.Div("COMPOSIÇÃO DE RECEITAS", style={'fontFamily':"'Space Mono',monospace",'fontSize':'9px','color':'#3ddc84','textTransform':'uppercase','letterSpacing':'3px','marginBottom':'3px'}),
                        html.Div("Vendas + Outras Receitas = Receita Total", style={'fontFamily':"'Playfair Display',serif",'fontSize':'15px','fontWeight':'700','color':'#e8ede9'}),
                    ]),
                    dcc.Graph(id='fig-receitas', config={'displayModeBar':False}, style={'height':'340px'}),
                ]),
            ]),
        ])

    elif tab == 'tab-txn':
        status_color = {'Confirmado':'#3ddc84','Processado':'#c9a84c','Pendente':'#ff5c5c'}
        cat_color    = {'Receita':'#3ddc84','Despesa':'#ff5c5c'}
        rows = []
        for t in TRANSACOES_MOCK:
            sc = status_color.get(t['status'], '#8fa894')
            cc = cat_color.get(t['categoria'], '#8fa894')
            r_rgb = '61,220,132' if cc == '#3ddc84' else '255,92,92'
            s_rgb = {'Confirmado':'61,220,132','Processado':'201,168,76','Pendente':'255,92,92'}.get(t['status'],'100,100,100')
            rows.append(html.Tr(style={'borderBottom':'1px solid rgba(255,255,255,0.04)'}, children=[
                html.Td(t['id'],        style={'padding':'13px 20px','fontFamily':"'Space Mono',monospace",'fontSize':'10px','color':'#4d5e52'}),
                html.Td(t['data'],      style={'padding':'13px 16px','fontFamily':"'Space Mono',monospace",'fontSize':'11px','color':'#8fa894'}),
                html.Td(t['descricao'], style={'padding':'13px 16px','fontSize':'13px','color':'#e8ede9'}),
                html.Td(html.Span(t['categoria'], style={'padding':'3px 10px','borderRadius':'4px','fontSize':'11px','fontWeight':'700','color':cc,'background':f'rgba({r_rgb},0.1)'}), style={'padding':'13px 16px'}),
                html.Td(t['valor'],     style={'padding':'13px 16px','fontFamily':"'Space Mono',monospace",'fontSize':'12px','color':'#e8ede9','fontWeight':'600','textAlign':'right'}),
                html.Td(t['variacao'],  style={'padding':'13px 16px','fontFamily':"'Space Mono',monospace",'fontSize':'11px','color':'#c9a84c','textAlign':'center'}),
                html.Td(html.Span(t['status'], style={'padding':'3px 10px','borderRadius':'4px','fontSize':'10px','fontWeight':'700','color':sc,'background':f'rgba({s_rgb},0.1)','fontFamily':"'Space Mono',monospace",'letterSpacing':'1px'}), style={'padding':'13px 20px'}),
            ]))
        return html.Div([
            html.Div(className='section-header', children=[html.Div([html.Div("Histórico Financeiro", className='section-eyebrow'), html.Div("Transações Recentes", className='section-title')])]),
            html.Div(className='table-card', children=[
                html.Div(className='table-header', children=[html.Div("Lançamentos — Dezembro 2025", className='table-header-title'), html.Div(f"{len(TRANSACOES_MOCK)} registros", style={'fontFamily':"'Space Mono',monospace",'fontSize':'10px','color':'#4d5e52'})]),
                html.Table(style={'width':'100%','borderCollapse':'collapse'}, children=[
                    html.Thead(children=[html.Tr(style={'background':'#1a1f1c','borderBottom':'2px solid rgba(255,255,255,0.06)'}, children=[
                        html.Th(h, style={'padding':'11px 20px' if i in (0,6) else '11px 16px','textAlign':'right' if i==4 else 'left','fontFamily':"'Space Mono',monospace",'fontSize':'9px','color':'#4d5e52','letterSpacing':'2px','textTransform':'uppercase'})
                        for i,h in enumerate(["ID","Data","Descrição","Categoria","Valor","Var.","Status"])])]),
                    html.Tbody(rows),
                ]),
            ]),
        ])

    return html.Div("Selecione uma aba.", style={'color':'#8fa894','padding':'40px'})

# MODAL SALVAR DRE
@app.callback(
    Output('modal-save-overlay','style'),
    Input('nav-save-dre','n_clicks'), Input('btn-save-close','n_clicks'),
    Input('btn-save-cancel','n_clicks'), Input('btn-save-confirm','n_clicks'),
    prevent_initial_call=True,
)
def toggle_save_modal(*_):
    return {'display':'block'} if ctx.triggered_id == 'nav-save-dre' else {'display':'none'}

@app.callback(
    Output('dre-log','data'), Output('save-status','children'),
    Input('btn-save-confirm','n_clicks'),
    State('save-mes','value'), State('save-ano','value'),
    State('dre-store','data'), State('dre-log','data'),
    prevent_initial_call=True,
)
def save_dre_to_log(n, mes, ano, dre_data, log):
    if not n: raise dash.exceptions.PreventUpdate
    if not mes or not ano: return dash.no_update, "⚠ Selecione mês e ano."
    if not dre_data: return dash.no_update, "⚠ Nenhum dado no DRE."
    log = log or {}
    key = f"{ano}-{mes}"
    MESES_PT = {'01':'Janeiro','02':'Fevereiro','03':'Março','04':'Abril','05':'Maio','06':'Junho',
                '07':'Julho','08':'Agosto','09':'Setembro','10':'Outubro','11':'Novembro','12':'Dezembro'}
    label = f"{MESES_PT.get(mes,mes)} / {ano}"
    log[key] = {**dre_data, '_label':label, '_key':key, '_saved_at':datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}
    return log, f"✓ DRE de {label} salvo!"

@app.callback(
    Output('modal-history-overlay','style'),
    Input('nav-history','n_clicks'), Input('btn-history-close','n_clicks'),
    prevent_initial_call=True,
)
def toggle_history_modal(*_):
    return {'display':'block'} if ctx.triggered_id == 'nav-history' else {'display':'none'}

@app.callback(
    Output('history-list','children'),
    Input('modal-history-overlay','style'), Input('search-mes','value'), Input('search-ano','value'),
    State('dre-log','data'), prevent_initial_call=True,
)
def render_history_list(style, mes_filter, ano_filter, log):
    if not log:
        return html.Div("Nenhum DRE salvo ainda.", style={'color':'#4d5e52','fontFamily':"'Space Mono',monospace",'fontSize':'11px','padding':'20px','textAlign':'center'})
    rows = []
    for key in sorted(log.keys(), reverse=True):
        entry = log[key]; kano, kmes = key.split('-')
        if mes_filter and kmes != mes_filter: continue
        if ano_filter and kano != ano_filter: continue
        rv  = (entry.get('venda_vista',0) or 0) + (entry.get('venda_prazo',0) or 0)
        rv += sum(entry.get(k,0) or 0 for k in ['or_juros','or_alug','or_outras','or_fundos','or_bonif'])
        cp  = sum(entry.get(k,0) or 0 for k in ['c_bebidas','c_bolos','c_carnes','c_cigarros','c_conv','c_estcoz','c_picoles','c_salgados','c_insumos','c_embal','c_recarga','c_bichos','c_buffet'])
        cf  = sum(entry.get(k,0) or 0 for k in ['cf_energia','cf_tel','cf_agua','cf_iptu','cf_sal','cf_enc','cf_imp','cf_seg','cf_diar'])
        cv  = sum(entry.get(k,0) or 0 for k in ['cv_sist','cv_terc','cv_exped','cv_honor','cv_manut','cv_viag','cv_taxas','cv_unif','cv_limp','cv_alug','cv_caixa','cv_mora','cv_banco','cv_cart','cv_brindes','cv_utens','cv_veic','cv_segpred','cv_gerais','cv_gas','cv_estoque','cv_comiss'])
        lucro = rv - cp - cf - cv
        def rbr(val): return f"R$ {val:,.2f}".replace(',','X').replace('.', ',').replace('X','.')
        rows.append(html.Div(style={'background':'#131714','border':'1px solid rgba(255,255,255,0.07)','borderRadius':'10px','padding':'16px 20px','marginBottom':'10px'}, children=[
            html.Div(style={'display':'flex','justifyContent':'space-between','alignItems':'center','marginBottom':'12px'}, children=[
                html.Div([html.Div(entry.get('_label',key), style={'fontFamily':"'Playfair Display',serif",'fontSize':'16px','fontWeight':'700','color':'#e8ede9'}),
                          html.Div(f"Salvo em {entry.get('_saved_at','')}", style={'fontFamily':"'Space Mono',monospace",'fontSize':'9px','color':'#4d5e52','marginTop':'2px'})]),
                html.Button("📂 Carregar", id={'type':'btn-load-dre','index':key}, style={'background':'rgba(61,220,132,0.1)','border':'1px solid rgba(61,220,132,0.3)','color':'#3ddc84','padding':'7px 14px','borderRadius':'6px','cursor':'pointer','fontFamily':"'Space Mono',monospace",'fontSize':'10px','fontWeight':'700'}),
            ]),
            html.Div(style={'display':'grid','gridTemplateColumns':'1fr 1fr 1fr 1fr','gap':'8px'}, children=[
                html.Div(style={'background':'rgba(61,220,132,0.06)','borderRadius':'6px','padding':'8px 12px'}, children=[html.Div("Receita",style={'fontFamily':"'Space Mono',monospace",'fontSize':'8px','color':'#4d5e52','textTransform':'uppercase','letterSpacing':'2px'}), html.Div(rbr(rv),style={'fontFamily':"'Space Mono',monospace",'fontSize':'11px','color':'#3ddc84','marginTop':'3px','fontWeight':'700'})]),
                html.Div(style={'background':'rgba(255,92,92,0.05)','borderRadius':'6px','padding':'8px 12px'}, children=[html.Div("Compras",style={'fontFamily':"'Space Mono',monospace",'fontSize':'8px','color':'#4d5e52','textTransform':'uppercase','letterSpacing':'2px'}), html.Div(rbr(cp),style={'fontFamily':"'Space Mono',monospace",'fontSize':'11px','color':'#ff5c5c','marginTop':'3px','fontWeight':'700'})]),
                html.Div(style={'background':'rgba(255,92,92,0.05)','borderRadius':'6px','padding':'8px 12px'}, children=[html.Div("Custos",style={'fontFamily':"'Space Mono',monospace",'fontSize':'8px','color':'#4d5e52','textTransform':'uppercase','letterSpacing':'2px'}), html.Div(rbr(cf+cv),style={'fontFamily':"'Space Mono',monospace",'fontSize':'11px','color':'#ff5c5c','marginTop':'3px','fontWeight':'700'})]),
                html.Div(style={'background':'rgba(201,168,76,0.07)','borderRadius':'6px','padding':'8px 12px'}, children=[html.Div("Lucro",style={'fontFamily':"'Space Mono',monospace",'fontSize':'8px','color':'#4d5e52','textTransform':'uppercase','letterSpacing':'2px'}), html.Div(rbr(lucro),style={'fontFamily':"'Space Mono',monospace",'fontSize':'11px','color':'#3ddc84' if lucro >= 0 else '#ff5c5c','marginTop':'3px','fontWeight':'700'})]),
            ]),
        ]))
    if not rows:
        return html.Div("Nenhum DRE encontrado para o filtro.", style={'color':'#4d5e52','fontFamily':"'Space Mono',monospace",'fontSize':'11px','padding':'20px','textAlign':'center'})
    return rows

@app.callback(
    Output('csv-store', 'data', allow_duplicate=True),
    Output('modal-history-overlay', 'style', allow_duplicate=True),
    Output('toast', 'children', allow_duplicate=True),
    Output('toast', 'style', allow_duplicate=True),
    Input({'type':'btn-load-dre','index':dash.ALL}, 'n_clicks'),
    State('dre-log','data'), prevent_initial_call=True,
)
def load_dre_from_history(n_clicks_list, log):
    if not any(n for n in n_clicks_list if n): raise dash.exceptions.PreventUpdate
    triggered = ctx.triggered_id
    if not triggered: raise dash.exceptions.PreventUpdate
    key = triggered['index']
    entry = (log or {}).get(key)
    if not entry: raise dash.exceptions.PreventUpdate
    label = entry.get('_label', key)
    toast = html.Div([html.Span("📂",style={'fontSize':'16px'}), html.Span(f" DRE de {label} carregado", style={'fontFamily':"'Space Mono',monospace",'fontSize':'11px'})], className='toast')
    clean = {k: v for k, v in entry.items() if not k.startswith('_')}
    return clean, {'display':'none'}, toast, {'display':'block'}

# ══════════════════════════════════════════════════════════════════════════════
# CALLBACKS — CONCILIAÇÃO BANCÁRIA
# ══════════════════════════════════════════════════════════════════════════════

@app.callback(
    Output('extrato-store', 'data'),
    Output('extrato-status-text', 'children'),
    Output('extrato-status-text', 'className'),
    Output('toast', 'children', allow_duplicate=True),
    Output('toast', 'style', allow_duplicate=True),
    Input('upload-extrato', 'contents'),
    State('upload-extrato', 'filename'),
    prevent_initial_call=True,
)
def processar_upload_extrato(contents, filename):
    if not contents:
        raise dash.exceptions.PreventUpdate

    _, content_string = contents.split(',')
    raw_bytes = base64.b64decode(content_string)

    try:
        df = processar_extrato_excel(raw_bytes)
    except Exception as e:
        toast = html.Div([
            html.Span("⚠", style={'color':'#ff5c5c','fontWeight':'700','fontSize':'16px'}),
            html.Span(f" Erro ao processar extrato: {str(e)}", style={'fontFamily':"'Space Mono',monospace",'fontSize':'11px'}),
        ], className='toast')
        return dash.no_update, "ERRO NO ARQUIVO", "upload-status", toast, {'display':'block'}

    # Serializa para JSON
    df_json = df.to_json(date_format='iso', orient='records', force_ascii=False)
    n = len(df)
    toast = html.Div([
        html.Span("✓", style={'color':'#3ddc84','fontWeight':'700','fontSize':'16px'}),
        html.Span(f" {filename} processado · {n} transações classificadas", style={'fontFamily':"'Space Mono',monospace",'fontSize':'11px'}),
    ], className='toast')
    return df_json, f"✓ {n} TRANSAÇÕES", "upload-status loaded", toast, {'display':'block'}


@app.callback(
    Output('conc-kpis',   'children'),
    Output('conc-charts', 'children'),
    Output('conc-table',  'children'),
    Input('extrato-store', 'data'),
    Input('filtro-tipo',   'value'),
    Input('filtro-es',     'value'),
    Input('filtro-busca',  'value'),
    prevent_initial_call=False,
)
def render_conciliacao(extrato_json, filtro_tipo, filtro_es, filtro_busca):
    # ── Estado vazio ──────────────────────────────────────────────────────────
    if not extrato_json:
        empty = html.Div(
            style={'textAlign':'center','padding':'60px 20px','color':'#4d5e52'},
            children=[
                html.Div("🏦", style={'fontSize':'48px','marginBottom':'16px'}),
                html.Div("Nenhum extrato importado ainda",
                         style={'fontFamily':"'Playfair Display',serif",'fontSize':'20px','color':'#8fa894','marginBottom':'8px'}),
                html.Div("Importe um arquivo Excel de extrato bancário acima para visualizar a conciliação.",
                         style={'fontFamily':"'Space Mono',monospace",'fontSize':'11px','letterSpacing':'1px'}),
            ]
        )
        return html.Div(), html.Div(), empty

    import json
    df = pd.DataFrame(json.loads(extrato_json))
    df['Valor (R$)'] = pd.to_numeric(df['Valor (R$)'], errors='coerce').fillna(0)

    # ── Filtros ───────────────────────────────────────────────────────────────
    df_f = df.copy()
    if filtro_tipo and filtro_tipo != 'TODOS':
        df_f = df_f[df_f['Tipo'] == filtro_tipo]
    if filtro_es and filtro_es != 'TODOS':
        df_f = df_f[df_f['Entrada/Saída'] == filtro_es]
    if filtro_busca:
        df_f = df_f[df_f['Descrição'].str.upper().str.contains(filtro_busca.upper(), na=False)]

    # ── KPIs ──────────────────────────────────────────────────────────────────
    total_entrada = df[df['Valor (R$)'] > 0]['Valor (R$)'].sum()
    total_saida   = df[df['Valor (R$)'] < 0]['Valor (R$)'].sum()
    saldo_periodo = total_entrada + total_saida
    n_transacoes  = len(df)
    n_tipos       = df['Tipo'].nunique()

    def rbr(v):
        return f"R$ {abs(v):,.2f}".replace(',','X').replace('.', ',').replace('X','.')

    kpis = html.Div(className='kpi-grid-5', children=[
        kpi_card("Total Entradas",   rbr(total_entrada),  f"{len(df[df['Valor (R$)']>0])} transações", "CRÉDITO", "up",   "green"),
        kpi_card("Total Saídas",     rbr(abs(total_saida)),f"{len(df[df['Valor (R$)']<0])} transações","DÉBITO",  "down", "red"),
        kpi_card("Saldo do Período", rbr(saldo_periodo),  "Crédito − Débito",
                 "POSITIVO" if saldo_periodo >= 0 else "NEGATIVO",
                 "up" if saldo_periodo >= 0 else "down",
                 "green" if saldo_periodo >= 0 else "red"),
        kpi_card("Nº de Transações", str(n_transacoes),  "Total no extrato",  "TOTAL",  "blue", "blue"),
        kpi_card("Tipos Detectados", str(n_tipos),        "Categorias únicas", "TIPOS",  "wait", "gold"),
    ])

    # ── Gráficos ──────────────────────────────────────────────────────────────
    # Gráfico 1: Entradas vs Saídas por Tipo
    tipo_group = df.groupby('Tipo')['Valor (R$)'].agg(
        Entradas=lambda x: x[x > 0].sum(),
        Saidas=lambda x: abs(x[x < 0].sum())
    ).reset_index()
    tipo_group = tipo_group[(tipo_group['Entradas'] > 0) | (tipo_group['Saidas'] > 0)]
    tipo_group = tipo_group.sort_values('Entradas', ascending=False)

    fig_tipos = go.Figure()
    if not tipo_group.empty:
        fig_tipos.add_trace(go.Bar(name='Entradas', x=tipo_group['Tipo'], y=tipo_group['Entradas'],
            marker_color='rgba(61,220,132,0.75)', marker_line_width=0,
            hovertemplate='<b>%{x}</b><br>Entrada: R$ %{y:,.2f}<extra></extra>'))
        fig_tipos.add_trace(go.Bar(name='Saídas', x=tipo_group['Tipo'], y=tipo_group['Saidas'],
            marker_color='rgba(255,92,92,0.65)', marker_line_width=0,
            hovertemplate='<b>%{x}</b><br>Saída: R$ %{y:,.2f}<extra></extra>'))
    layout1 = dict(CHART_LAYOUT)
    layout1['barmode']    = 'group'
    layout1['margin']     = dict(l=16, r=16, t=16, b=100)
    layout1['showlegend'] = True
    layout1['xaxis']      = dict(gridcolor='rgba(255,255,255,0.04)', tickfont=dict(size=10,color='#e8ede9',family='Sora'), tickangle=-30)
    fig_tipos.update_layout(**layout1)

    # Gráfico 2: Pizza de volume por tipo (entradas)
    entrada_tipos = df[df['Valor (R$)'] > 0].groupby('Tipo')['Valor (R$)'].sum().reset_index()
    entrada_tipos = entrada_tipos[entrada_tipos['Valor (R$)'] > 0].sort_values('Valor (R$)', ascending=False)
    colors_pie = [TYPE_COLORS.get(t, '#8fa894') for t in entrada_tipos['Tipo']]
    fig_pizza = go.Figure()
    if not entrada_tipos.empty:
        fig_pizza = go.Figure(go.Pie(
            labels=entrada_tipos['Tipo'], values=entrada_tipos['Valor (R$)'],
            hole=0.5, marker=dict(colors=colors_pie, line=dict(color='#0d0f0e', width=2)),
            textfont=dict(family='Space Mono', size=9, color='#0d0f0e'),
            hovertemplate='<b>%{label}</b><br>R$ %{value:,.2f}<br>%{percent}<extra></extra>',
        ))
    layout2 = dict(CHART_LAYOUT)
    layout2['margin'] = dict(l=16, r=16, t=16, b=60)
    fig_pizza.update_layout(**layout2)

    charts = html.Div(className='charts-grid-2', style={'marginBottom':'20px'}, children=[
        html.Div(style={'background':'#131714','border':'1px solid rgba(255,255,255,0.07)','borderRadius':'10px','overflow':'hidden'}, children=[
            html.Div(style={'padding':'14px 20px 10px','borderBottom':'1px solid rgba(255,255,255,0.06)','background':'#1a1f1c'}, children=[
                html.Div("ENTRADAS vs SAÍDAS", style={'fontFamily':"'Space Mono',monospace",'fontSize':'9px','color':'#3ddc84','textTransform':'uppercase','letterSpacing':'3px','marginBottom':'2px'}),
                html.Div("Volume por Tipo de Transação", style={'fontFamily':"'Playfair Display',serif",'fontSize':'14px','fontWeight':'700','color':'#e8ede9'}),
            ]),
            dcc.Graph(figure=fig_tipos, config={'displayModeBar':False}, style={'height':'280px'}),
        ]),
        html.Div(style={'background':'#131714','border':'1px solid rgba(255,255,255,0.07)','borderRadius':'10px','overflow':'hidden'}, children=[
            html.Div(style={'padding':'14px 20px 10px','borderBottom':'1px solid rgba(255,255,255,0.06)','background':'#1a1f1c'}, children=[
                html.Div("COMPOSIÇÃO DE ENTRADAS", style={'fontFamily':"'Space Mono',monospace",'fontSize':'9px','color':'#3ddc84','textTransform':'uppercase','letterSpacing':'3px','marginBottom':'2px'}),
                html.Div("Distribuição por Tipo (% do total)", style={'fontFamily':"'Playfair Display',serif",'fontSize':'14px','fontWeight':'700','color':'#e8ede9'}),
            ]),
            dcc.Graph(figure=fig_pizza, config={'displayModeBar':False}, style={'height':'280px'}),
        ]),
    ])

    # ── Tabela ────────────────────────────────────────────────────────────────
    rows_html = []
    for _, row in df_f.iterrows():
        val  = float(row['Valor (R$)'])
        es   = row['Entrada/Saída']
        tipo = row['Tipo']
        tc   = TYPE_COLORS.get(tipo, '#8fa894')
        tc_rgb = tc.lstrip('#')
        r_int  = int(tc_rgb[0:2], 16)
        g_int  = int(tc_rgb[2:4], 16)
        b_int  = int(tc_rgb[4:6], 16)
        val_fmt = f"R$ {abs(val):,.2f}".replace(',','X').replace('.', ',').replace('X','.')
        saldo_val = row.get('Saldo (R$)', None)
        try:
            saldo_fmt = f"R$ {float(saldo_val):,.2f}".replace(',','X').replace('.', ',').replace('X','.') if saldo_val is not None else '—'
        except:
            saldo_fmt = str(saldo_val) if saldo_val else '—'

        rows_html.append(html.Tr([
            html.Td(str(row['Data']),         style={'padding':'9px 14px','color':'#8fa894','fontFamily':"'Space Mono',monospace",'fontSize':'10px'}),
            html.Td(html.Span("⬆ " + es if es == "ENTRADA" else "⬇ " + es,
                               className='es-entrada' if es == "ENTRADA" else 'es-saida'),
                    style={'padding':'9px 14px'}),
            html.Td(html.Span(tipo, className='tipo-badge',
                               style={'color':tc,'background':f'rgba({r_int},{g_int},{b_int},0.12)','border':f'1px solid rgba({r_int},{g_int},{b_int},0.3)'}),
                    style={'padding':'9px 14px'}),
            html.Td(str(row['Descrição'])[:60] + ('…' if len(str(row['Descrição'])) > 60 else ''),
                    style={'padding':'9px 14px','color':'#8fa894','fontSize':'11px','maxWidth':'260px','overflow':'hidden'}),
            html.Td(str(row.get('Documento','')),
                    style={'padding':'9px 14px','fontFamily':"'Space Mono',monospace",'fontSize':'9px','color':'#4d5e52'}),
            html.Td(str(row['Categoria']),
                    style={'padding':'9px 14px','fontSize':'10px','color':'#8fa894','fontFamily':"'Space Mono',monospace"}),
            html.Td(val_fmt,
                    className='val-entrada' if val > 0 else 'val-saida',
                    style={'padding':'9px 14px','textAlign':'right'}),
            html.Td(saldo_fmt,
                    className='val-saldo',
                    style={'padding':'9px 14px','textAlign':'right'}),
        ]))

    tabela = html.Div(className='table-card', children=[
        html.Div(className='table-header', children=[
            html.Div(f"Transações — {len(df_f)} registros exibidos de {len(df)} total", className='table-header-title'),
            html.Div(style={'display':'flex','gap':'8px'}, children=[
                html.Span(f"✓ Filtros ativos" if (filtro_tipo != 'TODOS' or filtro_es != 'TODOS' or filtro_busca) else "",
                          style={'fontFamily':"'Space Mono',monospace",'fontSize':'10px','color':'#3ddc84'}),
            ]),
        ]),
        html.Div(className='conc-table-wrap', children=[
            html.Table(className='conc-table', children=[
                html.Thead(html.Tr([
                    html.Th(h) for h in ["Data","E/S","Tipo","Descrição","Documento","Categoria","Valor","Saldo"]
                ])),
                html.Tbody(rows_html if rows_html else [
                    html.Tr(html.Td("Nenhuma transação encontrada com os filtros aplicados.",
                                    colSpan=8, style={'textAlign':'center','padding':'40px','color':'#4d5e52',
                                                      'fontFamily':"'Space Mono',monospace",'fontSize':'11px'}))
                ]),
            ]),
        ]),
    ])

    return kpis, charts, tabela


@app.callback(
    Output('download-conc-excel', 'data'),
    Output('toast', 'children', allow_duplicate=True),
    Output('toast', 'style', allow_duplicate=True),
    Input('btn-export-conc', 'n_clicks'),
    State('extrato-store', 'data'),
    prevent_initial_call=True,
)
def exportar_conciliacao(n, extrato_json):
    if not n or not extrato_json:
        raise dash.exceptions.PreventUpdate
    import json
    df = pd.DataFrame(json.loads(extrato_json))
    df['Valor (R$)'] = pd.to_numeric(df['Valor (R$)'], errors='coerce').fillna(0)
    excel_bytes = gerar_excel_conciliacao(df)
    fname = f"Conciliacao_Bancaria_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    toast = html.Div([
        html.Span("✓", style={'color':'#3ddc84','fontWeight':'700','fontSize':'16px'}),
        html.Span(f" Conciliação exportada — {fname}", style={'fontFamily':"'Space Mono',monospace",'fontSize':'11px'}),
    ], className='toast')
    return dcc.send_bytes(excel_bytes, fname), toast, {'display':'block'}


app.clientside_callback(
    """
    function(style) {
        if (style && style.display === 'block') {
            setTimeout(function() {
                var el = document.getElementById('toast');
                if (el) el.style.display = 'none';
            }, 3500);
        }
        return window.dash_clientside.no_update;
    }
    """,
    Output('toast', 'style', allow_duplicate=True),
    Input('toast', 'style'),
    prevent_initial_call=True,
)

# ══════════════════════════════════════════════════════════════════════════════
# CALLBACKS — INDICADORES REAIS DO DRE
# ══════════════════════════════════════════════════════════════════════════════

def _kpi_inner(label, value_str, sub, badge, badge_cls, value_cls=""):
    """Retorna o conteúdo interno de um kpi-card."""
    return [
        html.Div(className='kpi-top', children=[
            html.Div(label, className='kpi-label'),
            html.Div(badge, className=f'kpi-badge {badge_cls}'),
        ]),
        html.Div(value_str, className=f'kpi-value {value_cls}'),
        html.Div(sub, className='kpi-sub'),
        html.Div(className='kpi-bar'),
    ]

def _brl(v):
    if v == 0: return "R$ 0,00"
    if abs(v) >= 1_000_000:
        return f"R$ {v/1e6:.2f}M".replace('.', ',')
    if abs(v) >= 1_000:
        return f"R$ {v/1000:.1f}K".replace('.', ',')
    return f"R$ {v:,.2f}".replace(',','X').replace('.', ',').replace('X','.')

def _pct(v, total):
    if not total: return "—"
    return f"{v/total*100:.1f}%"

@app.callback(
    Output('ind-receita-total',    'children'),
    Output('ind-margem-contrib',   'children'),
    Output('ind-lucro-bruto',      'children'),
    Output('ind-lucro-liquido',    'children'),
    Output('ind-total-compras',    'children'),
    Output('ind-custos-fixos',     'children'),
    Output('ind-custos-variaveis', 'children'),
    Output('ind-margem-pct',       'children'),
    Input('dre-store', 'data'),
    prevent_initial_call=False,
)
def update_indicadores(d):
    if not d:
        vazio = _kpi_inner("—", "—", "Preencha o DRE", "AGUARD.", "wait")
        return [vazio] * 8

    def v(k): return float(d.get(k, 0) or 0)

    receita_vendas = v('venda_vista') + v('venda_prazo')
    outras_rec     = v('or_juros') + v('or_alug') + v('or_outras') + v('or_fundos') + v('or_bonif')
    receita_total  = receita_vendas + outras_rec

    total_compras = sum(v(k) for k in [
        'c_bebidas','c_bolos','c_carnes','c_cigarros','c_conv','c_estcoz',
        'c_picoles','c_salgados','c_insumos','c_embal','c_recarga','c_bichos','c_buffet'])

    total_cf = sum(v(k) for k in [
        'cf_energia','cf_tel','cf_agua','cf_iptu','cf_sal',
        'cf_enc','cf_imp','cf_seg','cf_diar'])

    total_cv = sum(v(k) for k in [
        'cv_sist','cv_terc','cv_exped','cv_honor','cv_manut','cv_viag',
        'cv_taxas','cv_unif','cv_limp','cv_alug','cv_caixa','cv_mora',
        'cv_banco','cv_cart','cv_brindes','cv_utens','cv_veic',
        'cv_segpred','cv_gerais','cv_gas','cv_estoque','cv_comiss'])

    margem_contrib = receita_total - total_compras
    lucro_bruto    = margem_contrib - total_cf
    lucro_liquido  = lucro_bruto - total_cv

    # Classificação da margem %
    if receita_total > 0:
        m_pct = lucro_liquido / receita_total * 100
        if m_pct >= 30:    m_badge, m_cls = "EXCELENTE", "up"
        elif m_pct >= 15:  m_badge, m_cls = "BOM", "up"
        elif m_pct >= 0:   m_badge, m_cls = "ATENÇÃO", "wait"
        else:              m_badge, m_cls = "PREJUÍZO", "down"
    else:
        m_pct, m_badge, m_cls = 0, "AGUARD.", "wait"

    def lucro_badge(val):
        if val > 0:   return "LUCRO", "up"
        elif val < 0: return "PREJUÍZO", "down"
        else:         return "ZERO", "wait"

    lb, lcls = lucro_badge(lucro_liquido)
    gb, gcls = lucro_badge(lucro_bruto)
    mb, mcls = lucro_badge(margem_contrib)

    RT = receita_total if receita_total else 1

    return [
        _kpi_inner("Receita Total",        _brl(receita_total),   f"Vendas + Outras receitas", "RECEITA", "up"),
        _kpi_inner("Margem de Contribuição",_brl(margem_contrib), _pct(margem_contrib, RT) + " da receita", mb, mcls, "green" if margem_contrib >= 0 else "red"),
        _kpi_inner("Lucro Bruto",          _brl(lucro_bruto),     _pct(lucro_bruto, RT) + " da receita", gb, gcls, "green" if lucro_bruto >= 0 else "red"),
        _kpi_inner("Resultado Líquido",    _brl(lucro_liquido),   _pct(lucro_liquido, RT) + " da receita", lb, lcls, "green" if lucro_liquido >= 0 else "red"),
        _kpi_inner("Total Compras",        _brl(total_compras),   _pct(total_compras, RT) + " da receita", "CUSTO", "down"),
        _kpi_inner("Custos Fixos",         _brl(total_cf),        _pct(total_cf, RT) + " da receita", "FIXO", "wait"),
        _kpi_inner("Custos Variáveis",     _brl(total_cv),        _pct(total_cv, RT) + " da receita", "VARIÁVEL", "wait"),
        _kpi_inner("Margem Líquida %",     f"{m_pct:.1f}%",       "Sobre Receita Total", m_badge, m_cls, "gold" if m_pct >= 15 else ("red" if m_pct < 0 else "")),
    ]


@app.callback(
    Output('ind-cards-wrapper', 'style'),
    Output('btn-toggle-ind', 'children'),
    Input('btn-toggle-ind', 'n_clicks'),
    prevent_initial_call=False,
)
def toggle_indicadores(n):
    if n and n % 2 == 1:
        return {'display': 'none'}, "⊞ Mostrar Indicadores"
    return {'display': 'block'}, "⊟ Ocultar Indicadores"


# ── Callbacks dos botões da Home ──────────────────────────────────────────────
@app.callback(
    Output('active-tab',    'data',  allow_duplicate=True),
    Output('nav-ind',       'className', allow_duplicate=True),
    Output('nav-charts',    'className', allow_duplicate=True),
    Output('nav-dre',       'className', allow_duplicate=True),
    Output('nav-txn',       'className', allow_duplicate=True),
    Output('nav-conc',      'className', allow_duplicate=True),
    Output('topbar-title',  'children', allow_duplicate=True),
    Output('topbar-subtitle','children',allow_duplicate=True),
    Output('main-content',  'style',  allow_duplicate=True),
    Output('dre-tab-wrapper','style', allow_duplicate=True),
    Output('conc-tab-wrapper','style',allow_duplicate=True),
    Input('home-btn-dre',    'n_clicks'),
    Input('home-btn-charts', 'n_clicks'),
    Input('home-btn-conc',   'n_clicks'),
    Input('home-btn-txn',    'n_clicks'),
    Input('home-btn-export', 'n_clicks'),
    prevent_initial_call=True,
)
def home_buttons(n_dre, n_charts, n_conc, n_txn, n_export):
    show = {'display':'block'}; hide = {'display':'none'}
    btn_map = {
        'home-btn-dre':    ('tab-dre',    'nav-dre',    'DRE Completo',         'Demonstração do Resultado do Exercício · Panelas do Guapo'),
        'home-btn-charts': ('tab-charts', 'nav-charts', 'Análise Gráfica',      'Visualização de dados financeiros'),
        'home-btn-conc':   ('tab-conc',   'nav-conc',   'Conciliação Bancária', 'Extrato bancário · Classificação automática'),
        'home-btn-txn':    ('tab-txn',    'nav-txn',    'Transações Recentes',  'Histórico de entradas e saídas'),
        'home-btn-export': ('tab-dre',    'nav-dre',    'DRE Completo',         'Use as ferramentas na sidebar para exportar'),
    }
    tid = ctx.triggered_id
    if not tid or tid not in btn_map:
        raise dash.exceptions.PreventUpdate
    tab, nav_active, title, sub = btn_map[tid]
    nav_ids = ['nav-ind','nav-charts','nav-dre','nav-txn','nav-conc']
    classes = ['nav-item active' if n == nav_active else 'nav-item' for n in nav_ids]
    is_dre  = tab == 'tab-dre'
    is_conc = tab == 'tab-conc'
    return (tab, *classes, title, sub,
            hide if (is_dre or is_conc) else show,
            show if is_dre  else hide,
            show if is_conc else hide)


    port = int(os.environ.get("PORT", 8050))
    app.run_server(host='0.0.0.0', port=port, debug=False)
